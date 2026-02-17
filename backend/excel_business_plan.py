"""
Excel Report Generator for Business Plan
Generates comprehensive multi-sheet Excel that mirrors the entire web application.
THIS IS A FULLY FORMULA-DRIVEN WORKBOOK - all calculations happen in Excel.
"""

import io
from datetime import datetime
from typing import Dict, Any, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# === STYLE CONSTANTS ===
COLORS = {
    'primary': '1E3A5F',
    'input_blue': '0000FF',
    'formula_black': '000000',
    'link_green': '008000',
    'white': 'FFFFFF',
    'input_bg': 'FFFFCC',
    'header_bg': '1E3A5F',
    'light_gray': 'F5F5F5',
    'border_gray': 'CCCCCC',
}

THIN_BORDER = Border(
    left=Side(style='thin', color=COLORS['border_gray']),
    right=Side(style='thin', color=COLORS['border_gray']),
    top=Side(style='thin', color=COLORS['border_gray']),
    bottom=Side(style='thin', color=COLORS['border_gray'])
)

HEADER_FILL = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type='solid')
INPUT_FILL = PatternFill(start_color=COLORS['input_bg'], end_color=COLORS['input_bg'], fill_type='solid')
LIGHT_FILL = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type='solid')

HEADER_FONT = Font(name='Calibri', size=11, bold=True, color=COLORS['white'])
INPUT_FONT = Font(name='Calibri', size=11, color=COLORS['input_blue'])
FORMULA_FONT = Font(name='Calibri', size=11, color=COLORS['formula_black'])
LINK_FONT = Font(name='Calibri', size=11, color=COLORS['link_green'])
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color=COLORS['primary'])
SECTION_FONT = Font(name='Calibri', size=12, bold=True, color=COLORS['primary'])
BOLD_FONT = Font(name='Calibri', size=11, bold=True)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')


class BusinessPlanExcelGenerator:
    """
    Generates a FULLY FORMULA-DRIVEN Business Plan Excel.
    ALL calculations happen in Excel - no pre-calculated Python values.
    """

    def __init__(self, lot_key: str, business_plan: Dict[str, Any], costs: Dict[str, float],
                 clean_team_cost: float, base_amount: float, is_rti: bool = False,
                 quota_lutech: float = 1.0, scenarios: Optional[List[Dict[str, Any]]] = None,
                 tow_breakdown: Optional[Dict[str, Any]] = None,
                 profile_rates: Optional[Dict[str, float]] = None,
                 intervals: Optional[List[Dict[str, Any]]] = None,
                 lutech_breakdown: Optional[Dict[str, Any]] = None):

        self.lot_key = lot_key
        self.bp = business_plan or {}
        self.costs = costs or {}
        self.clean_team_cost = clean_team_cost
        self.base_amount = base_amount
        self.is_rti = is_rti
        self.quota_lutech = quota_lutech
        self.scenarios = scenarios or []
        self.tow_breakdown = tow_breakdown or {}
        self.lutech_breakdown = lutech_breakdown or {}
        self.profile_rates = profile_rates or {}
        self.intervals = intervals or []

        self.wb = Workbook()

        # Named ranges for cross-sheet references
        self.named_ranges = {}

    def generate(self) -> io.BytesIO:
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Create sheets in order - PARAMS FIRST so other sheets can reference them
        self._create_params_sheet()           # 1. All input parameters
        self._create_lutech_catalog_sheet()   # 2. Lutech profiles with rates
        self._create_tow_config_sheet()       # 3. TOW configuration
        self._create_team_sheet()             # 4. Team (Poste profiles, FTE, TOW allocation)
        self._create_volume_adj_sheet()       # 5. Rettifica Volumi (riduzione per periodo)
        self._create_mapping_sheet()          # 6. Profile mapping Poste → Lutech
        self._create_subcontract_sheet()      # 7. Subcontract config
        self._create_tow_analysis_sheet()     # 8. ANALISI TOW (Margine per TOW) - NEW!
        self._create_cost_calc_sheet()        # 9. Cost calculations (ALL FORMULAS)
        self._create_pl_sheet()               # 10. P&L (ALL FORMULAS)
        self._create_scenarios_sheet()        # 11. Scenarios
        self._create_offer_sheet()            # 12. Offer scheme (ALL FORMULAS)

        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _style_input_cell(self, cell):
        cell.fill = INPUT_FILL
        cell.font = INPUT_FONT
        cell.border = THIN_BORDER

    def _style_formula_cell(self, cell):
        cell.font = FORMULA_FONT
        cell.border = THIN_BORDER

    def _style_link_cell(self, cell):
        cell.font = LINK_FONT
        cell.border = THIN_BORDER

    def _add_header_row(self, ws, row: int, headers: List[str], start_col: int = 1):
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=start_col + i, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    # ========== SHEET 1: PARAMETRI (Central input sheet) ==========
    def _create_params_sheet(self):
        ws = self.wb.create_sheet("PARAMETRI", 0)
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 45

        row = 2

        # === HEADER PROFESSIONALE ===
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = 'SIMULATORE GARA POSTE'
        ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color=COLORS['primary'])
        ws[f'A{row}'].alignment = CENTER
        row += 1

        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = f'BUSINESS PLAN - {self.lot_key}'
        ws[f'A{row}'].font = Font(name='Calibri', size=20, bold=True, color=COLORS['primary'])
        ws[f'A{row}'].alignment = CENTER
        row += 1

        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = f'Generato il {datetime.now().strftime("%d/%m/%Y alle %H:%M")}'
        ws[f'A{row}'].font = Font(name='Calibri', size=10, italic=True, color='666666')
        ws[f'A{row}'].alignment = CENTER
        row += 2

        # === CONTENUTO PARAMETRI ===
        row += 1
        ws['A' + str(row)] = "PARAMETRI GARA"
        ws['A' + str(row)].font = SECTION_FONT
        row += 1

        # Base d'asta
        ws['A' + str(row)] = "Base d'asta (€)"
        ws['B' + str(row)] = self.base_amount
        ws['B' + str(row)].number_format = '#,##0'
        self._style_input_cell(ws['B' + str(row)])
        ws['C' + str(row)] = "Importo totale della gara"
        ws['C' + str(row)].font = Font(italic=True, color='666666')
        self.named_ranges['BASE_ASTA'] = f"PARAMETRI!$B${row}"
        row += 1

        # RTI
        ws['A' + str(row)] = "RTI Attivo (1=Sì, 0=No)"
        ws['B' + str(row)] = 1 if self.is_rti else 0
        self._style_input_cell(ws['B' + str(row)])
        self.named_ranges['RTI_ATTIVO'] = f"PARAMETRI!$B${row}"
        row += 1

        # Quota Lutech
        ws['A' + str(row)] = "Quota Lutech RTI"
        ws['B' + str(row)] = self.quota_lutech
        ws['B' + str(row)].number_format = '0.0%'
        self._style_input_cell(ws['B' + str(row)])
        self.named_ranges['QUOTA_LUTECH'] = f"PARAMETRI!$B${row}"
        row += 1

        # Base effettiva (FORMULA)
        ws['A' + str(row)] = "Base Effettiva Lutech (€)"
        ws['B' + str(row)] = f"=IF({self.named_ranges['RTI_ATTIVO']}=1,{self.named_ranges['BASE_ASTA']}*{self.named_ranges['QUOTA_LUTECH']},{self.named_ranges['BASE_ASTA']})"
        ws['B' + str(row)].number_format = '#,##0'
        self._style_formula_cell(ws['B' + str(row)])
        ws['C' + str(row)] = "Formula: Base × Quota (se RTI)"
        ws['C' + str(row)].font = Font(italic=True, color='666666')
        self.named_ranges['BASE_EFFETTIVA'] = f"PARAMETRI!$B${row}"
        row += 2

        # PARAMETRI TEMPORALI
        ws['A' + str(row)] = "PARAMETRI TEMPORALI"
        ws['A' + str(row)].font = SECTION_FONT
        row += 1

        ws['A' + str(row)] = "Durata Contratto (mesi)"
        ws['B' + str(row)] = self.bp.get('duration_months', 36)
        self._style_input_cell(ws['B' + str(row)])
        self.named_ranges['DURATA_MESI'] = f"PARAMETRI!$B${row}"
        row += 1

        ws['A' + str(row)] = "Giorni/Anno per FTE"
        ws['B' + str(row)] = self.bp.get('days_per_fte', 220)
        self._style_input_cell(ws['B' + str(row)])
        self.named_ranges['GG_ANNO'] = f"PARAMETRI!$B${row}"
        row += 1

        ws['A' + str(row)] = "Tariffa Default (€/gg)"
        ws['B' + str(row)] = self.bp.get('default_daily_rate', 250)
        ws['B' + str(row)].number_format = '#,##0'
        self._style_input_cell(ws['B' + str(row)])
        self.named_ranges['TARIFFA_DEFAULT'] = f"PARAMETRI!$B${row}"
        row += 2

        # FATTORI DI COSTO
        ws['A' + str(row)] = "FATTORI DI COSTO"
        ws['A' + str(row)].font = SECTION_FONT
        row += 1

        gov_pct = self.bp.get('governance_pct', 0.04)
        if gov_pct > 1:
            gov_pct = gov_pct / 100
        ws['A' + str(row)] = "Governance %"
        ws['B' + str(row)] = gov_pct
        ws['B' + str(row)].number_format = '0.0%'
        self._style_input_cell(ws['B' + str(row)])
        self.named_ranges['GOVERNANCE_PCT'] = f"PARAMETRI!$B${row}"
        row += 1

        risk_pct = self.bp.get('risk_contingency_pct', 0.03)
        if risk_pct > 1:
            risk_pct = risk_pct / 100
        ws['A' + str(row)] = "Risk Contingency %"
        ws['B' + str(row)] = risk_pct
        ws['B' + str(row)].number_format = '0.0%'
        self._style_input_cell(ws['B' + str(row)])
        self.named_ranges['RISK_PCT'] = f"PARAMETRI!$B${row}"
        row += 1

        reuse = self.bp.get('reuse_factor', 0)
        if reuse > 1:
            reuse = reuse / 100
        ws['A' + str(row)] = "Reuse Factor %"
        ws['B' + str(row)] = reuse
        ws['B' + str(row)].number_format = '0.0%'
        self._style_input_cell(ws['B' + str(row)])
        self.named_ranges['REUSE_FACTOR'] = f"PARAMETRI!$B${row}"
        row += 2

        # OFFERTA
        ws['A' + str(row)] = "OFFERTA"
        ws['A' + str(row)].font = SECTION_FONT
        row += 1

        ws['A' + str(row)] = "Sconto Offerta %"
        ws['B' + str(row)] = 0.05
        ws['B' + str(row)].number_format = '0.0%'
        self._style_input_cell(ws['B' + str(row)])
        self.named_ranges['SCONTO'] = f"PARAMETRI!$B${row}"
        row += 1

        ws['A' + str(row)] = "Margine Target %"
        ws['B' + str(row)] = 0.15
        ws['B' + str(row)].number_format = '0.0%'
        self._style_input_cell(ws['B' + str(row)])
        self.named_ranges['MARGINE_TARGET'] = f"PARAMETRI!$B${row}"
        row += 1

        # Revenue (FORMULA)
        ws['A' + str(row)] = "Revenue (€)"
        ws['B' + str(row)] = f"={self.named_ranges['BASE_EFFETTIVA']}*(1-{self.named_ranges['SCONTO']})"
        ws['B' + str(row)].number_format = '#,##0'
        self._style_formula_cell(ws['B' + str(row)])
        ws['B' + str(row)].fill = LIGHT_FILL
        ws['C' + str(row)] = "Formula: Base Effettiva × (1 - Sconto)"
        ws['C' + str(row)].font = Font(italic=True, color='666666')
        self.named_ranges['REVENUE'] = f"PARAMETRI!$B${row}"

        # === FOOTER PROFESSIONALE ===
        row += 3
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = '─' * 60
        ws[f'A{row}'].font = Font(size=8, color='666666')
        ws[f'A{row}'].alignment = CENTER
        row += 1

        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = 'https://simulator-poste.c-6dc1be8.kyma.ondemand.com'
        ws[f'A{row}'].font = Font(name='Calibri', size=9, color='2563EB', underline='single')
        ws[f'A{row}'].alignment = CENTER
        ws[f'A{row}'].hyperlink = 'https://simulator-poste.c-6dc1be8.kyma.ondemand.com'
        row += 1

        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = 'Sviluppato da Gabriele Rendina'
        ws[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color='666666')
        ws[f'A{row}'].alignment = CENTER

    # ========== SHEET 2: CATALOGO LUTECH ==========
    def _create_lutech_catalog_sheet(self):
        ws = self.wb.create_sheet("CATALOGO_LUTECH")
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15

        ws['A1'] = "CATALOGO PROFILI LUTECH"
        ws['A1'].font = TITLE_FONT
        ws['A2'] = "Le tariffe in questa tabella sono usate per calcolare i costi del team."
        ws['A2'].font = Font(italic=True, size=10, color='666666')

        row = 4
        headers = ['ID Profilo', 'Label', 'Tariffa (€/gg)']
        self._add_header_row(ws, row, headers)
        row += 1
        data_start = row

        # Populate from profile_rates
        for full_id, rate in sorted(self.profile_rates.items()):
            parts = full_id.split(':')
            label = parts[1] if len(parts) > 1 else full_id

            ws.cell(row=row, column=1, value=full_id).border = THIN_BORDER
            ws.cell(row=row, column=2, value=label).border = THIN_BORDER

            cell_rate = ws.cell(row=row, column=3, value=rate)
            cell_rate.number_format = '#,##0'
            self._style_input_cell(cell_rate)

            row += 1

        self.named_ranges['CATALOGO_START'] = data_start
        self.named_ranges['CATALOGO_END'] = row - 1
        self.named_ranges['CATALOGO_RANGE'] = f"CATALOGO_LUTECH!$A${data_start}:$C${row-1}"

    # ========== SHEET 3: CONFIGURAZIONE TOW ==========
    def _create_tow_config_sheet(self):
        ws = self.wb.create_sheet("CONFIG_TOW")
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12

        ws['A1'] = "CONFIGURAZIONE TOW"
        ws['A1'].font = TITLE_FONT

        row = 3
        headers = ['TOW ID', 'Descrizione', 'Tipo', 'Num Task', 'Peso %']
        self._add_header_row(ws, row, headers)
        row += 1
        data_start = row

        tows = self.bp.get('tows', [])

        # Create dropdown for type
        type_validation = DataValidation(
            type="list",
            formula1='"task,corpo,consumo"',
            allow_blank=False,
            showDropDown=False
        )
        type_validation.error = "Seleziona task, corpo o consumo"
        type_validation.errorTitle = "Tipo non valido"
        type_validation.prompt = "Scegli il tipo di TOW"
        type_validation.promptTitle = "Tipo TOW"

        for tow in tows:
            ws.cell(row=row, column=1, value=tow.get('tow_id', tow.get('id', ''))).border = THIN_BORDER
            ws.cell(row=row, column=2, value=tow.get('label', '')).border = THIN_BORDER

            cell_type = ws.cell(row=row, column=3, value=tow.get('type', 'task'))
            self._style_input_cell(cell_type)
            cell_type.alignment = CENTER

            cell_tasks = ws.cell(row=row, column=4, value=tow.get('num_tasks', 0) or 0)
            self._style_input_cell(cell_tasks)
            cell_tasks.alignment = CENTER

            cell_weight = ws.cell(row=row, column=5, value=(tow.get('weight_pct', 0) or 0) / 100)
            cell_weight.number_format = '0.0%'
            self._style_input_cell(cell_weight)
            cell_weight.alignment = CENTER

            row += 1

        # Add validation to type column
        if tows:
            type_validation.add(f'C{data_start}:C{row-1}')
            ws.add_data_validation(type_validation)

        self.named_ranges['TOW_START'] = data_start
        self.named_ranges['TOW_END'] = row - 1
        self.named_ranges['TOW_RANGE'] = f"CONFIG_TOW!$A${data_start}:$E${row-1}"

    # ========== SHEET 4: TEAM COMPOSITION (No Lutech mapping - just Poste profiles) ==========
    def _create_team_sheet(self):
        ws = self.wb.create_sheet("TEAM")

        # Get TOWs for dynamic columns
        tows = self.bp.get('tows') or self.bp.get('tow_config') or []
        tow_ids = [t.get('tow_id', '') for t in tows]
        num_tows = len(tow_ids)

        # Column widths
        ws.column_dimensions['A'].width = 28  # Profilo
        ws.column_dimensions['B'].width = 14  # Seniority
        ws.column_dimensions['C'].width = 13  # FTE Base
        ws.column_dimensions['D'].width = 13  # GG/Anno
        # Dynamic TOW columns start at E
        for i in range(num_tows):
            col_letter = get_column_letter(5 + i)
            ws.column_dimensions[col_letter].width = 13
        # FTE Eff and GG Totali after TOW columns
        fte_eff_col = 5 + num_tows  # FTE Effettivo
        gg_tot_col = 6 + num_tows   # GG Totali
        ws.column_dimensions[get_column_letter(fte_eff_col)].width = 14
        ws.column_dimensions[get_column_letter(gg_tot_col)].width = 14

        # Freeze header row and profile column
        ws.freeze_panes = 'B5'

        ws['A1'] = "COMPOSIZIONE TEAM POSTE"
        ws['A1'].font = TITLE_FONT
        ws['A2'] = "FTE, Allocazione TOW e Fattori Riduzione sono INPUT. FTE Eff e GG sono FORMULE."
        ws['A2'].font = Font(italic=True, size=10, color='666666')

        row = 4
        # Build headers: Profilo, Seniority, FTE Base, GG/Anno, [TOW1 %, TOW2 %, ...], FTE Eff, GG Totali
        headers = ['Profilo', 'Seniority', 'FTE Base', 'GG/Anno']
        for tow in tows:
            headers.append(f"{tow.get('tow_id', tow.get('id', ''))} %")
        headers.extend(['FTE Eff.', 'GG Totali'])
        self._add_header_row(ws, row, headers)
        row += 1
        data_start = row

        team = self.bp.get('team_composition', []) or self.bp.get('team', [])
        volume_adj = self.bp.get('volume_adjustments', {})
        # Get period factors (use first period or default)
        periods = volume_adj.get('periods', [{}])
        first_period = periods[0] if periods else {}
        by_profile = first_period.get('by_profile', {})
        by_tow = first_period.get('by_tow', {})

        # Seniority dropdown
        seniority_validation = DataValidation(
            type="list",
            formula1='"jr,mid,sr,expert"',
            allow_blank=False,
            showDropDown=False
        )

        for member in team:
            profile_id = member.get('profile_id', member.get('label', 'Unknown'))
            profile_label = member.get('label', profile_id)
            tow_allocation = member.get('tow_allocation', {})

            # Column A: Profilo
            ws.cell(row=row, column=1, value=profile_label).border = THIN_BORDER

            # Column B: Seniority (input)
            cell_sen = ws.cell(row=row, column=2, value=member.get('seniority', 'mid'))
            self._style_input_cell(cell_sen)
            cell_sen.alignment = CENTER

            # Column C: FTE Base (input)
            cell_fte = ws.cell(row=row, column=3, value=float(member.get('fte', 0)))
            cell_fte.number_format = '0.00'
            self._style_input_cell(cell_fte)
            cell_fte.alignment = CENTER

            # Column D: GG/Anno (calculated from FTE × 220)
            cell_gg_year = ws.cell(row=row, column=4)
            cell_gg_year.value = f"=C{row}*{self.named_ranges['GG_ANNO']}"
            cell_gg_year.number_format = '#,##0'
            self._style_formula_cell(cell_gg_year)
            cell_gg_year.alignment = CENTER

            # Columns E+: TOW allocation % (input)
            for i, tow in enumerate(tows):
                tow_id = tow.get('tow_id', tow.get('id', ''))
                alloc_pct = tow_allocation.get(tow_id, 0) / 100 if tow_allocation.get(tow_id, 0) > 0 else 0
                col = 5 + i
                cell_alloc = ws.cell(row=row, column=col, value=alloc_pct)
                cell_alloc.number_format = '0%'
                self._style_input_cell(cell_alloc)
                cell_alloc.alignment = CENTER

            # FTE Effettivo: FORMULA = FTE Base × Fattore Profilo × (1-Reuse) × Fattore TOW pesato
            # Fattore TOW pesato = Σ (Alloc% × FattoreTOW) / Σ Alloc% (calcolato manualmente in formula)
            # Per semplicità in Excel, applichiamo solo Fattore Profilo e Reuse (i TOW factors sono in RETTIFICA)
            profile_factor = by_profile.get(profile_id, 1.0)
            cell_fte_eff = ws.cell(row=row, column=fte_eff_col)
            # Formula: FTE × ProfileFactor × (1 - Reuse)
            cell_fte_eff.value = f"=C{row}*{profile_factor}*(1-{self.named_ranges['REUSE_FACTOR']})"
            cell_fte_eff.number_format = '0.00'
            self._style_formula_cell(cell_fte_eff)
            cell_fte_eff.alignment = CENTER

            # GG Totali: FORMULA = FTE Eff × GG/Anno × (Durata/12)
            cell_gg_tot = ws.cell(row=row, column=gg_tot_col)
            fte_eff_letter = get_column_letter(fte_eff_col)
            cell_gg_tot.value = f"={fte_eff_letter}{row}*{self.named_ranges['GG_ANNO']}*({self.named_ranges['DURATA_MESI']}/12)"
            cell_gg_tot.number_format = '#,##0'
            self._style_formula_cell(cell_gg_tot)
            cell_gg_tot.alignment = CENTER

            row += 1

        # Add seniority validation
        if team:
            seniority_validation.add(f'B{data_start}:B{row-1}')
            ws.add_data_validation(seniority_validation)

        # TOTALS row
        if team:
            ws.cell(row=row, column=1, value="TOTALE").font = BOLD_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER

            # Sum FTE Base
            ws.cell(row=row, column=3, value=f"=SUM(C{data_start}:C{row-1})")
            ws.cell(row=row, column=3).number_format = '0.00'
            ws.cell(row=row, column=3).font = BOLD_FONT
            ws.cell(row=row, column=3).border = THIN_BORDER

            # Sum GG/Anno
            ws.cell(row=row, column=4, value=f"=SUM(D{data_start}:D{row-1})")
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=4).font = BOLD_FONT
            ws.cell(row=row, column=4).border = THIN_BORDER

            # Sum FTE Eff
            fte_eff_letter = get_column_letter(fte_eff_col)
            ws.cell(row=row, column=fte_eff_col, value=f"=SUM({fte_eff_letter}{data_start}:{fte_eff_letter}{row-1})")
            ws.cell(row=row, column=fte_eff_col).number_format = '0.00'
            ws.cell(row=row, column=fte_eff_col).font = BOLD_FONT
            ws.cell(row=row, column=fte_eff_col).border = THIN_BORDER
            ws.cell(row=row, column=fte_eff_col).fill = LIGHT_FILL

            # Sum GG Totali
            gg_tot_letter = get_column_letter(gg_tot_col)
            ws.cell(row=row, column=gg_tot_col, value=f"=SUM({gg_tot_letter}{data_start}:{gg_tot_letter}{row-1})")
            ws.cell(row=row, column=gg_tot_col).number_format = '#,##0'
            ws.cell(row=row, column=gg_tot_col).font = BOLD_FONT
            ws.cell(row=row, column=gg_tot_col).border = THIN_BORDER
            ws.cell(row=row, column=gg_tot_col).fill = LIGHT_FILL

            self.named_ranges['TEAM_FTE_BASE'] = f"TEAM!$C${row}"
            self.named_ranges['TEAM_FTE_EFF'] = f"TEAM!${fte_eff_letter}${row}"
            self.named_ranges['TEAM_GG'] = f"TEAM!${gg_tot_letter}${row}"

        # Delta row (risparmio FTE)
        row += 1
        ws.cell(row=row, column=1, value="RISPARMIO FTE").font = Font(italic=True, color='008000')
        ws.cell(row=row, column=fte_eff_col, value=f"=C{row-1}-{get_column_letter(fte_eff_col)}{row-1}")
        ws.cell(row=row, column=fte_eff_col).number_format = '0.00'
        ws.cell(row=row, column=fte_eff_col).font = Font(italic=True, color='008000')
        ws.cell(row=row, column=fte_eff_col + 1, value=f"=(C{row-1}-{get_column_letter(fte_eff_col)}{row-1})/C{row-1}")
        ws.cell(row=row, column=fte_eff_col + 1).number_format = '0.0%'
        ws.cell(row=row, column=fte_eff_col + 1).font = Font(italic=True, color='008000')

    # ========== SHEET 5: RETTIFICA VOLUMI (Time-phased adjustments) ==========
    def _create_volume_adj_sheet(self):
        ws = self.wb.create_sheet("RETTIFICA_VOLUMI")
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 40

        ws['A1'] = "RETTIFICA VOLUMI PER PERIODO"
        ws['A1'].font = TITLE_FONT
        ws['A2'] = "Fattori di riduzione FTE per profilo e TOW, configurabili per periodo temporale."
        ws['A2'].font = Font(italic=True, size=10, color='666666')

        volume_adj = self.bp.get('volume_adjustments', {})
        periods = volume_adj.get('periods', [])
        team = self.bp.get('team_composition', []) or self.bp.get('team', [])
        tows = self.bp.get('tows') or self.bp.get('tow_config') or []
        duration_months = self.bp.get('duration_months', 36)

        row = 4

        # If no periods, create a default one
        if not periods:
            periods = [{
                'month_start': 1,
                'month_end': duration_months,
                'by_profile': {},
                'by_tow': {}
            }]

        for p_idx, period in enumerate(periods):
            month_start = period.get('month_start', 1)
            month_end = period.get('month_end', duration_months)
            by_profile = period.get('by_profile', {})
            by_tow = period.get('by_tow', {})

            # Period header
            ws.cell(row=row, column=1, value=f"PERIODO {p_idx + 1}: Mesi {month_start} - {month_end}")
            ws.cell(row=row, column=1).font = SECTION_FONT
            row += 1

            # Month range inputs
            ws.cell(row=row, column=1, value="Mese Inizio:")
            cell_start = ws.cell(row=row, column=2, value=month_start)
            self._style_input_cell(cell_start)
            cell_start.alignment = CENTER

            ws.cell(row=row, column=3, value="Mese Fine:")
            cell_end = ws.cell(row=row, column=4, value=month_end)
            self._style_input_cell(cell_end)
            cell_end.alignment = CENTER
            row += 2

            # Profile reduction factors
            ws.cell(row=row, column=1, value="RIDUZIONE FTE PER PROFILO")
            ws.cell(row=row, column=1).font = Font(bold=True, color='8B008B')  # Purple
            row += 1

            headers = ['Profilo', 'Fattore %', '', 'Effetto', 'Note']
            self._add_header_row(ws, row, headers)
            row += 1

            for member in team:
                profile_id = member.get('profile_id', member.get('label', ''))
                profile_label = member.get('label', profile_id)
                fte = float(member.get('fte', 0))
                factor = by_profile.get(profile_id, 1.0)

                ws.cell(row=row, column=1, value=profile_label).border = THIN_BORDER

                cell_factor = ws.cell(row=row, column=2, value=factor)
                cell_factor.number_format = '0%'
                self._style_input_cell(cell_factor)
                cell_factor.alignment = CENTER

                # Effect (calculated)
                fte_eff = fte * factor
                ws.cell(row=row, column=4, value=f"{fte:.1f} → {fte_eff:.1f} FTE")
                ws.cell(row=row, column=4).font = Font(color='008000' if factor < 1.0 else '666666')

                if factor < 1.0:
                    ws.cell(row=row, column=5, value=f"Riduzione {(1-factor)*100:.0f}%")
                    ws.cell(row=row, column=5).font = Font(italic=True, color='008000')

                row += 1

            row += 1

            # TOW reduction factors
            ws.cell(row=row, column=1, value="RIDUZIONE PER TOW")
            ws.cell(row=row, column=1).font = Font(bold=True, color='DAA520')  # Amber/Gold
            row += 1

            headers = ['TOW', 'Tipo', 'Fattore %', 'Effetto', 'Note']
            self._add_header_row(ws, row, headers)
            row += 1

            for tow in tows:
                tow_id = tow.get('tow_id', tow.get('id', ''))
                tow_type = tow.get('type', 'task')
                factor = by_tow.get(tow_id, 1.0)

                ws.cell(row=row, column=1, value=tow_id).border = THIN_BORDER
                ws.cell(row=row, column=2, value=tow_type).border = THIN_BORDER

                cell_factor = ws.cell(row=row, column=3, value=factor)
                cell_factor.number_format = '0%'
                self._style_input_cell(cell_factor)
                cell_factor.alignment = CENTER

                # Effect based on TOW type
                if tow_type == 'task':
                    num_tasks = tow.get('num_tasks', 0)
                    eff_tasks = int(num_tasks * factor)
                    ws.cell(row=row, column=4, value=f"{num_tasks} → {eff_tasks} task")
                elif tow_type == 'corpo':
                    dur = tow.get('duration_months', duration_months)
                    eff_dur = dur * factor
                    ws.cell(row=row, column=4, value=f"{dur} → {eff_dur:.1f} mesi")
                else:
                    ws.cell(row=row, column=4, value="N/A (consumo)")

                ws.cell(row=row, column=4).font = Font(color='008000' if factor < 1.0 else '666666')

                if factor < 1.0:
                    ws.cell(row=row, column=5, value=f"Riduzione {(1-factor)*100:.0f}%")
                    ws.cell(row=row, column=5).font = Font(italic=True, color='008000')

                row += 1

            row += 2  # Space before next period

        # Explanation box
        ws.cell(row=row, column=1, value="LOGICA DI CALCOLO:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        explanations = [
            "FTE Effettivo = FTE Base × Fattore Profilo × (1 - Reuse%) × Fattore TOW pesato",
            "Fattore TOW pesato = Σ (Allocazione% × Fattore TOW) / Σ Allocazione%",
            "Un fattore al 80% significa riduzione del 20% rispetto al valore base.",
            "I TOW 'a consumo' non prevedono rettifiche volume automatiche."
        ]
        for exp in explanations:
            ws.cell(row=row, column=1, value=f"• {exp}")
            ws.cell(row=row, column=1).font = Font(italic=True, color='666666')
            row += 1

    # ========== SHEET 6: MAPPING PROFILI (with cost calculation and VALIDATIONS) ==========
    def _create_mapping_sheet(self):
        ws = self.wb.create_sheet("MAPPING")
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 16
        ws.freeze_panes = 'A5'

        ws['A1'] = "MAPPING PROFILI POSTE → LUTECH E CALCOLO COSTO"
        ws['A1'].font = TITLE_FONT
        ws['A2'] = "⚠️ IMPORTANTE: Usare le dropdown per selezionare i profili corretti!"
        ws['A2'].font = Font(italic=True, size=10, color='CC0000')

        row = 4
        headers = ['Profilo Poste', 'Profilo Lutech', 'Mix %', 'Tariffa (€/gg)', 'GG', 'Costo (€)']
        self._add_header_row(ws, row, headers)
        row += 1
        data_start = row

        team = self.bp.get('team_composition', []) or self.bp.get('team', [])
        mappings = self.bp.get('profile_mapping', []) or self.bp.get('profile_mappings', {})

        # === VALIDATION 1: Profilo Poste dropdown (from team labels) ===
        poste_labels = [m.get('label', m.get('profile_id', '')) for m in team]
        poste_validation = None
        if poste_labels:
            poste_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(poste_labels[:50])}"',
                allow_blank=False,
                showDropDown=False
            )
            poste_validation.error = "Seleziona un profilo Poste dalla lista"
            poste_validation.errorTitle = "Profilo Poste non valido"
            poste_validation.prompt = "Scegli il profilo Poste da mappare"
            poste_validation.promptTitle = "Profilo Poste"
            poste_validation.showErrorMessage = True
            poste_validation.showInputMessage = True

        # === VALIDATION 2: Profilo Lutech dropdown (from catalog) ===
        lutech_ids = list(self.profile_rates.keys())
        # Also get from lutech_profiles in BP data
        lutech_profiles = self.bp.get('lutech_profiles', [])
        if lutech_profiles:
            lutech_ids = list(set(lutech_ids + [p.get('id', '') for p in lutech_profiles]))

        lutech_validation = None
        if lutech_ids:
            lutech_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(lutech_ids[:50])}"',
                allow_blank=False,
                showDropDown=False
            )
            lutech_validation.error = "Seleziona un profilo Lutech dal catalogo"
            lutech_validation.errorTitle = "Profilo Lutech non valido"
            lutech_validation.prompt = "Scegli il profilo Lutech dal catalogo"
            lutech_validation.promptTitle = "Profilo Lutech"
            lutech_validation.showErrorMessage = True
            lutech_validation.showInputMessage = True

        # === VALIDATION 3: Mix % (0-100%) ===
        mix_validation = DataValidation(
            type="decimal",
            operator="between",
            formula1="0",
            formula2="1",
            allow_blank=False,
            showDropDown=False
        )
        mix_validation.error = "Il valore deve essere tra 0% e 100%"
        mix_validation.errorTitle = "Mix % non valido"
        mix_validation.prompt = "Inserisci la percentuale di mix (0-100%)"
        mix_validation.promptTitle = "Mix %"
        mix_validation.showErrorMessage = True
        mix_validation.showInputMessage = True

        # If mappings is a list (new format), convert to dict
        if isinstance(mappings, list):
            mappings_dict = {}
            for m in mappings:
                poste_id = m.get('poste_profile_id', '')
                if poste_id not in mappings_dict:
                    mappings_dict[poste_id] = []
                mappings_dict[poste_id].append({
                    'month_start': 1,
                    'month_end': 36,
                    'mix': [{'lutech_profile': m.get('lutech_profile_id', ''), 'pct': 100}]
                })
            mappings = mappings_dict

        # Build mapping rows with cost calculation
        for member in team:
            profile_id = member.get('profile_id', member.get('label', ''))
            profile_label = member.get('label', profile_id)

            # Find mapping for this profile
            profile_mappings = mappings.get(profile_id, [])

            if not profile_mappings:
                # No mapping - create a default row with validation
                cell_poste = ws.cell(row=row, column=1, value=profile_label)
                cell_poste.border = THIN_BORDER
                self._style_input_cell(cell_poste)

                cell_lutech = ws.cell(row=row, column=2, value='')
                self._style_input_cell(cell_lutech)

                cell_mix = ws.cell(row=row, column=3, value=1.0)
                cell_mix.number_format = '0%'
                self._style_input_cell(cell_mix)
                cell_mix.alignment = CENTER

                # Tariffa: VLOOKUP from catalog
                cell_rate = ws.cell(row=row, column=4)
                cell_rate.value = f"=IFERROR(VLOOKUP(B{row},{self.named_ranges['CATALOGO_RANGE']},3,FALSE),{self.named_ranges['TARIFFA_DEFAULT']})"
                cell_rate.number_format = '#,##0'
                self._style_link_cell(cell_rate)

                # GG: proportional to FTE
                member_fte = float(member.get('fte', 0))
                total_fte = sum(float(m2.get('fte', 0)) for m2 in team)
                fte_share = member_fte / total_fte if total_fte > 0 else 0

                cell_gg = ws.cell(row=row, column=5)
                cell_gg.value = f"={self.named_ranges.get('TEAM_GG', '0')}*{fte_share:.4f}"
                cell_gg.number_format = '#,##0'
                self._style_formula_cell(cell_gg)

                # Costo = GG × Mix × Tariffa
                cell_cost = ws.cell(row=row, column=6)
                cell_cost.value = f"=E{row}*C{row}*D{row}"
                cell_cost.number_format = '#,##0'
                self._style_formula_cell(cell_cost)

                row += 1
            else:
                first = True
                for period in profile_mappings:
                    mix_list = period.get('mix', [])
                    if not mix_list:
                        mix_list = [{'lutech_profile': '', 'pct': 100}]

                    for m in mix_list:
                        cell_poste = ws.cell(row=row, column=1, value=profile_label if first else "")
                        cell_poste.border = THIN_BORDER
                        if first:
                            self._style_input_cell(cell_poste)

                        lutech_profile = m.get('lutech_profile', '')
                        cell_lutech = ws.cell(row=row, column=2, value=lutech_profile)
                        self._style_input_cell(cell_lutech)

                        pct = (m.get('pct', 100) or 100) / 100
                        cell_mix = ws.cell(row=row, column=3, value=pct)
                        cell_mix.number_format = '0%'
                        self._style_input_cell(cell_mix)
                        cell_mix.alignment = CENTER

                        # Tariffa: VLOOKUP from catalog
                        cell_rate = ws.cell(row=row, column=4)
                        cell_rate.value = f"=IFERROR(VLOOKUP(B{row},{self.named_ranges['CATALOGO_RANGE']},3,FALSE),{self.named_ranges['TARIFFA_DEFAULT']})"
                        cell_rate.number_format = '#,##0'
                        self._style_link_cell(cell_rate)

                        # GG: proportional to FTE
                        member_fte = float(member.get('fte', 0))
                        total_fte = sum(float(m2.get('fte', 0)) for m2 in team)
                        fte_share = member_fte / total_fte if total_fte > 0 else 0

                        cell_gg = ws.cell(row=row, column=5)
                        cell_gg.value = f"={self.named_ranges.get('TEAM_GG', '0')}*{fte_share:.4f}"
                        cell_gg.number_format = '#,##0'
                        self._style_formula_cell(cell_gg)

                        # Costo = GG × Mix × Tariffa
                        cell_cost = ws.cell(row=row, column=6)
                        cell_cost.value = f"=E{row}*C{row}*D{row}"
                        cell_cost.number_format = '#,##0'
                        self._style_formula_cell(cell_cost)

                        row += 1
                        first = False

        # === ADD ALL VALIDATIONS TO RANGES ===
        if row > data_start:
            if poste_validation:
                poste_validation.add(f'A{data_start}:A{row-1}')
                ws.add_data_validation(poste_validation)

            if lutech_validation:
                lutech_validation.add(f'B{data_start}:B{row-1}')
                ws.add_data_validation(lutech_validation)

            mix_validation.add(f'C{data_start}:C{row-1}')
            ws.add_data_validation(mix_validation)

        # TOTALE row
        if row > data_start:
            ws.cell(row=row, column=1, value="TOTALE COSTO TEAM").font = BOLD_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER

            ws.cell(row=row, column=5, value=f"=SUM(E{data_start}:E{row-1})")
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=5).font = BOLD_FONT
            ws.cell(row=row, column=5).border = THIN_BORDER

            ws.cell(row=row, column=6, value=f"=SUM(F{data_start}:F{row-1})")
            ws.cell(row=row, column=6).number_format = '#,##0'
            ws.cell(row=row, column=6).font = BOLD_FONT
            ws.cell(row=row, column=6).border = THIN_BORDER
            ws.cell(row=row, column=6).fill = LIGHT_FILL

            self.named_ranges['TEAM_COST'] = f"MAPPING!$F${row}"
            self.named_ranges['MAPPING_GG'] = f"MAPPING!$E${row}"
        else:
            # No team data - use fallback values
            ws.cell(row=row, column=1, value="(Nessun team definito)")
            ws.cell(row=row, column=1).font = Font(italic=True, color='999999')
            ws.cell(row=row, column=6, value=0)
            self.named_ranges['TEAM_COST'] = f"MAPPING!$F${row}"
            self.named_ranges['MAPPING_GG'] = "1"

        # Tariffa Media row
        row += 2
        ws.cell(row=row, column=1, value="TARIFFA MEDIA PONDERATA").font = BOLD_FONT
        ws.cell(row=row, column=4, value=f"=IFERROR({self.named_ranges.get('TEAM_COST', '0')}/{self.named_ranges.get('MAPPING_GG', '1')},0)")
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=4).font = BOLD_FONT
        ws.cell(row=row, column=4).fill = LIGHT_FILL
        ws.cell(row=row, column=4).border = THIN_BORDER
        self.named_ranges['TARIFFA_MEDIA'] = f"MAPPING!$D${row}"

        # Validation summary
        row += 2
        ws.cell(row=row, column=1, value="⚠️ VALIDAZIONI ATTIVE:").font = Font(bold=True, color='CC0000')
        row += 1
        ws.cell(row=row, column=1, value="• Colonna A: Solo profili Poste dal TEAM")
        ws.cell(row=row, column=1).font = Font(italic=True, color='666666')
        row += 1
        ws.cell(row=row, column=1, value="• Colonna B: Solo profili Lutech dal CATALOGO")
        ws.cell(row=row, column=1).font = Font(italic=True, color='666666')
        row += 1
        ws.cell(row=row, column=1, value="• Colonna C: Mix % deve essere tra 0% e 100%")
        ws.cell(row=row, column=1).font = Font(italic=True, color='666666')

    # ========== SHEET 7: SUBAPPALTO ==========
    def _create_subcontract_sheet(self):
        ws = self.wb.create_sheet("SUBAPPALTO")
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18

        ws['A1'] = "CONFIGURAZIONE SUBAPPALTO"
        ws['A1'].font = TITLE_FONT

        sub_cfg = self.bp.get('subcontract_config') or {}

        row = 3
        ws['A' + str(row)] = "Partner:"
        cell_partner = ws['B' + str(row)]
        cell_partner.value = sub_cfg.get('partner', '')
        self._style_input_cell(cell_partner)
        row += 1

        ws['A' + str(row)] = "Tariffa Media Partner (€/gg):"
        cell_rate = ws['B' + str(row)]
        cell_rate.value = sub_cfg.get('avg_daily_rate', 200) or 200
        cell_rate.number_format = '#,##0'
        self._style_input_cell(cell_rate)
        self.named_ranges['SUB_TARIFFA'] = f"SUBAPPALTO!$B${row}"
        row += 2

        ws['A' + str(row)] = "RIPARTIZIONE PER TOW"
        ws['A' + str(row)].font = SECTION_FONT
        row += 2

        headers = ['TOW ID', 'Quota Subappalto %', 'Costo Subappalto (€)']
        self._add_header_row(ws, row, headers)
        row += 1
        data_start = row

        tow_split = sub_cfg.get('tow_split') or {}
        tows = self.bp.get('tows') or self.bp.get('tow_config') or []

        for tow in tows:
            tow_id = tow.get('tow_id', tow.get('id', ''))
            split = (tow_split.get(tow_id, 0) or 0) / 100

            ws.cell(row=row, column=1, value=tow_id).border = THIN_BORDER

            cell_pct = ws.cell(row=row, column=2, value=split)
            cell_pct.number_format = '0.0%'
            self._style_input_cell(cell_pct)
            cell_pct.alignment = CENTER

            # Costo = Team Cost × Split %
            cell_cost = ws.cell(row=row, column=3)
            cell_cost.value = f"=B{row}*{self.named_ranges['TEAM_COST']}"
            cell_cost.number_format = '#,##0'
            self._style_formula_cell(cell_cost)

            row += 1

        # Totals
        if tows:
            ws.cell(row=row, column=1, value="TOTALE").font = BOLD_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER

            ws.cell(row=row, column=2, value=f"=SUM(B{data_start}:B{row-1})")
            ws.cell(row=row, column=2).number_format = '0.0%'
            ws.cell(row=row, column=2).font = BOLD_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER

            ws.cell(row=row, column=3, value=f"=SUM(C{data_start}:C{row-1})")
            ws.cell(row=row, column=3).number_format = '#,##0'
            ws.cell(row=row, column=3).font = BOLD_FONT
            ws.cell(row=row, column=3).border = THIN_BORDER
            ws.cell(row=row, column=3).fill = LIGHT_FILL

            self.named_ranges['SUB_TOTAL_PCT'] = f"SUBAPPALTO!$B${row}"
            self.named_ranges['SUB_COST'] = f"SUBAPPALTO!$C${row}"
        else:
            # No TOWs - create a placeholder cell with 0
            ws.cell(row=row, column=1, value="(Nessun TOW)")
            ws.cell(row=row, column=3, value=0)
            ws.cell(row=row, column=3).number_format = '#,##0'
            self.named_ranges['SUB_COST'] = f"SUBAPPALTO!$C${row}"

    # ========== SHEET 8: ANALISI TOW (Margine per TOW - Business Intelligence) ==========
    def _create_tow_analysis_sheet(self):
        """
        Analisi incrociata TOW × Team per supportare decisioni di business:
        1. Margine per TOW
        2. Allocazione Team per TOW (matrice)
        3. Concentrazione Senior vs Junior
        4. Rischi e Raccomandazioni
        """
        ws = self.wb.create_sheet("ANALISI_TOW")

        # Column widths
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 28
        ws.freeze_panes = 'C4'

        tows = self.bp.get('tows') or self.bp.get('tow_config') or []
        team = self.bp.get('team_composition', []) or self.bp.get('team', [])
        num_tows = len(tows)

        # ==================== SEZIONE 1: MARGINE PER TOW ====================
        ws['A1'] = "ANALISI BUSINESS PER TOW"
        ws['A1'].font = TITLE_FONT
        ws['A2'] = "Analisi incrociata Team × TOW per identificare profittabilità, concentrazione risorse e rischi."
        ws['A2'].font = Font(italic=True, size=10, color='666666')

        row = 4
        ws.cell(row=row, column=1, value="SEZIONE 1: MARGINE PER TOW").font = SECTION_FONT
        row += 2

        headers = ['TOW ID', 'Descrizione', 'Peso %', 'Ricavo (€)', 'Costo (€)', 'Margine (€)', 'Margine %', 'Status']
        self._add_header_row(ws, row, headers)
        row += 1
        margin_data_start = row

        # Calculate weights
        total_weight = sum(tow.get('weight_pct', 0) or 0 for tow in tows) or (len(tows) * 100)

        tow_rows = {}  # Track row for each TOW
        for tow in tows:
            tow_id = tow.get('tow_id', tow.get('id', ''))
            tow_label = tow.get('label', tow.get('tow_name', tow.get('description', '')))
            weight_pct = (tow.get('weight_pct', 0) or 0) / total_weight if total_weight > 0 else 1/len(tows)

            if weight_pct == 0:
                tow_cost = self.tow_breakdown.get(tow_id, 0)
                if isinstance(tow_cost, dict):
                    tow_cost = tow_cost.get('cost', 0)
                total_breakdown = sum((v.get('cost', 0) if isinstance(v, dict) else v) for v in self.tow_breakdown.values()) or 1
                weight_pct = tow_cost / total_breakdown if total_breakdown > 0 else 1/len(tows)

            tow_rows[tow_id] = row

            ws.cell(row=row, column=1, value=tow_id).border = THIN_BORDER
            ws.cell(row=row, column=2, value=tow_label).border = THIN_BORDER

            cell_weight = ws.cell(row=row, column=3, value=weight_pct)
            cell_weight.number_format = '0.0%'
            self._style_input_cell(cell_weight)

            # Ricavo
            cell_rev = ws.cell(row=row, column=4)
            cell_rev.value = f"={self.named_ranges.get('REVENUE', 'PARAMETRI!$B$22')}*C{row}"
            cell_rev.number_format = '#,##0'
            self._style_formula_cell(cell_rev)

            # Costo (will be linked to Section 2)
            cell_cost = ws.cell(row=row, column=5)
            cell_cost.value = f"={self.named_ranges.get('TEAM_COST', '0')}*C{row}"
            cell_cost.number_format = '#,##0'
            self._style_formula_cell(cell_cost)

            # Margine
            cell_margin = ws.cell(row=row, column=6)
            cell_margin.value = f"=D{row}-E{row}"
            cell_margin.number_format = '#,##0'
            self._style_formula_cell(cell_margin)

            # Margine %
            cell_margin_pct = ws.cell(row=row, column=7)
            cell_margin_pct.value = f"=IFERROR(F{row}/D{row},0)"
            cell_margin_pct.number_format = '0.0%'
            self._style_formula_cell(cell_margin_pct)

            # Status
            cell_status = ws.cell(row=row, column=8)
            cell_status.value = f'=IF(G{row}>=0.2,"✓ ALTO",IF(G{row}>=0.1,"○ OK",IF(G{row}>=0,"⚠ BASSO","✗ PERDITA")))'
            cell_status.border = THIN_BORDER

            row += 1

        margin_data_end = row - 1

        # Conditional formatting (only if there's data)
        if margin_data_end >= margin_data_start:
            from openpyxl.formatting.rule import CellIsRule
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            ws.conditional_formatting.add(f'G{margin_data_start}:G{margin_data_end}', CellIsRule(operator='greaterThanOrEqual', formula=['0.15'], fill=green_fill))
            ws.conditional_formatting.add(f'G{margin_data_start}:G{margin_data_end}', CellIsRule(operator='between', formula=['0', '0.15'], fill=yellow_fill))
            ws.conditional_formatting.add(f'G{margin_data_start}:G{margin_data_end}', CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))

        # Totals (only if there's data)
        if margin_data_end >= margin_data_start:
            ws.cell(row=row, column=1, value="TOTALE").font = BOLD_FONT
            for col in [3, 4, 5, 6]:
                ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{margin_data_start}:{get_column_letter(col)}{margin_data_end})")
                ws.cell(row=row, column=col).font = BOLD_FONT
                ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=3).number_format = '0.0%'
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=6).number_format = '#,##0'
            ws.cell(row=row, column=6).fill = LIGHT_FILL
            ws.cell(row=row, column=7, value=f"=IFERROR(F{row}/D{row},0)")
            ws.cell(row=row, column=7).number_format = '0.0%'
            ws.cell(row=row, column=7).font = BOLD_FONT
        else:
            ws.cell(row=row, column=1, value="(Nessun TOW configurato)").font = Font(italic=True, color='999999')

        # ==================== SEZIONE 2: MATRICE ALLOCAZIONE TEAM × TOW ====================
        row += 3
        ws.cell(row=row, column=1, value="SEZIONE 2: ALLOCAZIONE TEAM PER TOW").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        row += 2

        # Headers: Profilo, Seniority, FTE, Tariffa, then one column per TOW, then Total
        alloc_headers = ['Profilo', 'Seniority', 'FTE', 'Tariffa €/gg']
        for tow in tows:
            alloc_headers.append(f"{tow.get('tow_id', tow.get('id', ''))} %")
        alloc_headers.extend(['GG Totali', 'Costo Totale €'])
        self._add_header_row(ws, row, alloc_headers)
        row += 1
        alloc_data_start = row

        for member in team:
            profile_id = member.get('profile_id', '')
            profile_label = member.get('label', profile_id)
            seniority = member.get('seniority', 'mid')
            fte = float(member.get('fte', 0))
            tow_allocation = member.get('tow_allocation', {})

            # Get tariffa from mapping
            tariffa = self.profile_rates.get(member.get('lutech_profile_id', ''), 0)
            if tariffa == 0:
                # Try to find from profile_mapping (could be list or dict)
                pm_data = self.bp.get('profile_mapping', [])
                if isinstance(pm_data, dict):
                    # Dict format: {profile_id: [{'mix': [...]}]}
                    if profile_id in pm_data:
                        pm_entries = pm_data[profile_id]
                        if pm_entries and len(pm_entries) > 0:
                            mix_list = pm_entries[0].get('mix', [])
                            if mix_list:
                                lutech_id = mix_list[0].get('lutech_profile', '')
                                tariffa = self.profile_rates.get(lutech_id, 400)
                else:
                    # List format: [{'poste_profile_id': ..., 'lutech_profile_id': ...}]
                    for pm in pm_data:
                        if pm.get('poste_profile_id') == profile_id:
                            tariffa = self.profile_rates.get(pm.get('lutech_profile_id', ''), 400)
                            break
                if tariffa == 0:
                    tariffa = 400  # Default

            ws.cell(row=row, column=1, value=profile_label).border = THIN_BORDER
            ws.cell(row=row, column=2, value=seniority.upper()).border = THIN_BORDER
            ws.cell(row=row, column=2).alignment = CENTER

            cell_fte = ws.cell(row=row, column=3, value=fte)
            cell_fte.number_format = '0.00'
            self._style_input_cell(cell_fte)

            cell_tariffa = ws.cell(row=row, column=4, value=tariffa)
            cell_tariffa.number_format = '#,##0'
            self._style_input_cell(cell_tariffa)

            # TOW allocations
            col = 5
            for tow in tows:
                tow_id = tow.get('tow_id', tow.get('id', ''))
                alloc = tow_allocation.get(tow_id, 0) / 100 if tow_allocation.get(tow_id, 0) > 0 else 0
                cell_alloc = ws.cell(row=row, column=col, value=alloc)
                cell_alloc.number_format = '0%'
                self._style_input_cell(cell_alloc)
                cell_alloc.alignment = CENTER
                col += 1

            # GG Totali = FTE × GG/Anno × Durata
            cell_gg = ws.cell(row=row, column=col)
            cell_gg.value = f"=C{row}*{self.named_ranges.get('GG_ANNO', '220')}*({self.named_ranges.get('DURATA_MESI', '36')}/12)*(1-{self.named_ranges.get('REUSE_FACTOR', '0')})"
            cell_gg.number_format = '#,##0'
            self._style_formula_cell(cell_gg)
            col += 1

            # Costo Totale = GG × Tariffa
            cell_costo = ws.cell(row=row, column=col)
            cell_costo.value = f"={get_column_letter(col-1)}{row}*D{row}"
            cell_costo.number_format = '#,##0'
            self._style_formula_cell(cell_costo)

            row += 1

        alloc_data_end = row - 1
        gg_col = 5 + num_tows
        costo_col = 6 + num_tows

        # Totals row
        ws.cell(row=row, column=1, value="TOTALE").font = BOLD_FONT
        ws.cell(row=row, column=3, value=f"=SUM(C{alloc_data_start}:C{alloc_data_end})")
        ws.cell(row=row, column=3).font = BOLD_FONT
        ws.cell(row=row, column=3).number_format = '0.00'
        ws.cell(row=row, column=gg_col, value=f"=SUM({get_column_letter(gg_col)}{alloc_data_start}:{get_column_letter(gg_col)}{alloc_data_end})")
        ws.cell(row=row, column=gg_col).font = BOLD_FONT
        ws.cell(row=row, column=gg_col).number_format = '#,##0'
        ws.cell(row=row, column=costo_col, value=f"=SUM({get_column_letter(costo_col)}{alloc_data_start}:{get_column_letter(costo_col)}{alloc_data_end})")
        ws.cell(row=row, column=costo_col).font = BOLD_FONT
        ws.cell(row=row, column=costo_col).number_format = '#,##0'
        ws.cell(row=row, column=costo_col).fill = LIGHT_FILL

        # ==================== SEZIONE 3: CONCENTRAZIONE SENIOR vs JUNIOR ====================
        row += 3
        ws.cell(row=row, column=1, value="SEZIONE 3: CONCENTRAZIONE SENIOR vs JUNIOR PER TOW").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
        row += 2

        # Calculate senior/junior concentration per TOW
        senior_headers = ['TOW ID', 'Descrizione', 'FTE Senior', 'FTE Junior', '% Senior', '% Junior', 'Mix Index', 'Valutazione']
        self._add_header_row(ws, row, senior_headers)
        row += 1
        senior_data_start = row

        for tow in tows:
            tow_id = tow.get('tow_id', tow.get('id', ''))
            tow_label = tow.get('label', '')

            # Calculate weighted FTE per seniority for this TOW
            senior_fte = 0
            junior_fte = 0
            for member in team:
                fte = float(member.get('fte', 0))
                alloc = member.get('tow_allocation', {}).get(tow_id, 0) / 100
                seniority = member.get('seniority', 'mid')
                if seniority in ['sr', 'expert']:
                    senior_fte += fte * alloc
                else:
                    junior_fte += fte * alloc

            total_fte_tow = senior_fte + junior_fte

            ws.cell(row=row, column=1, value=tow_id).border = THIN_BORDER
            ws.cell(row=row, column=2, value=tow_label).border = THIN_BORDER

            ws.cell(row=row, column=3, value=round(senior_fte, 2)).border = THIN_BORDER
            ws.cell(row=row, column=3).number_format = '0.00'

            ws.cell(row=row, column=4, value=round(junior_fte, 2)).border = THIN_BORDER
            ws.cell(row=row, column=4).number_format = '0.00'

            sr_pct = senior_fte / total_fte_tow if total_fte_tow > 0 else 0
            jr_pct = junior_fte / total_fte_tow if total_fte_tow > 0 else 0

            cell_sr_pct = ws.cell(row=row, column=5, value=sr_pct)
            cell_sr_pct.number_format = '0%'
            cell_sr_pct.border = THIN_BORDER
            if sr_pct > 0.6:
                cell_sr_pct.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

            cell_jr_pct = ws.cell(row=row, column=6, value=jr_pct)
            cell_jr_pct.number_format = '0%'
            cell_jr_pct.border = THIN_BORDER

            # Mix Index (0 = all junior, 1 = all senior)
            ws.cell(row=row, column=7, value=sr_pct).border = THIN_BORDER
            ws.cell(row=row, column=7).number_format = '0.00'

            # Valutazione
            if sr_pct > 0.7:
                valutazione = "⚠ Troppi Senior"
            elif sr_pct < 0.2:
                valutazione = "⚠ Pochi Senior"
            else:
                valutazione = "✓ Mix Equilibrato"
            ws.cell(row=row, column=8, value=valutazione).border = THIN_BORDER

            row += 1

        # ==================== SEZIONE 4: RISCHI E RACCOMANDAZIONI ====================
        row += 2
        ws.cell(row=row, column=1, value="SEZIONE 4: ANALISI RISCHI E RACCOMANDAZIONI").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        row += 2

        # Key metrics
        ws.cell(row=row, column=1, value="📊 METRICHE CHIAVE:").font = Font(bold=True)
        row += 1

        ws.cell(row=row, column=1, value="TOW più profittevole:")
        ws.cell(row=row, column=2, value=f"=INDEX(A{margin_data_start}:A{margin_data_end},MATCH(MAX(G{margin_data_start}:G{margin_data_end}),G{margin_data_start}:G{margin_data_end},0))")
        ws.cell(row=row, column=2).font = Font(bold=True, color='008000')
        ws.cell(row=row, column=3, value=f"=MAX(G{margin_data_start}:G{margin_data_end})")
        ws.cell(row=row, column=3).number_format = '0.0%'
        row += 1

        ws.cell(row=row, column=1, value="TOW meno profittevole:")
        ws.cell(row=row, column=2, value=f"=INDEX(A{margin_data_start}:A{margin_data_end},MATCH(MIN(G{margin_data_start}:G{margin_data_end}),G{margin_data_start}:G{margin_data_end},0))")
        ws.cell(row=row, column=2).font = Font(bold=True, color='CC0000')
        ws.cell(row=row, column=3, value=f"=MIN(G{margin_data_start}:G{margin_data_end})")
        ws.cell(row=row, column=3).number_format = '0.0%'
        row += 1

        ws.cell(row=row, column=1, value="TOW in perdita:")
        ws.cell(row=row, column=2, value=f"=COUNTIF(G{margin_data_start}:G{margin_data_end},\"<0\")")
        ws.cell(row=row, column=2).font = BOLD_FONT
        row += 1

        ws.cell(row=row, column=1, value="Concentrazione costi top 1 TOW:")
        ws.cell(row=row, column=2, value=f"=MAX(E{margin_data_start}:E{margin_data_end})/SUM(E{margin_data_start}:E{margin_data_end})")
        ws.cell(row=row, column=2).number_format = '0.0%'
        row += 2

        # Risks
        ws.cell(row=row, column=1, value="⚠️ RISCHI IDENTIFICATI:").font = Font(bold=True, color='CC0000')
        row += 1

        ws.cell(row=row, column=1, value=f'=IF(COUNTIF(G{margin_data_start}:G{margin_data_end},"<0")>0,"• RISCHIO: Ci sono TOW in perdita! Rivedere allocazione.","")')
        ws.cell(row=row, column=1).font = Font(color='CC0000')
        row += 1

        ws.cell(row=row, column=1, value=f'=IF(MAX(E{margin_data_start}:E{margin_data_end})/SUM(E{margin_data_start}:E{margin_data_end})>0.5,"• RISCHIO: Alta concentrazione costi su un singolo TOW (>50%).","")')
        ws.cell(row=row, column=1).font = Font(color='CC0000')
        row += 1

        ws.cell(row=row, column=1, value=f'=IF(MIN(G{margin_data_start}:G{margin_data_end})<0.05,"• ATTENZIONE: Almeno un TOW ha margine <5%.","")')
        ws.cell(row=row, column=1).font = Font(color='CC6600')
        row += 2

        # Recommendations
        ws.cell(row=row, column=1, value="💡 RACCOMANDAZIONI:").font = Font(bold=True, color='008000')
        row += 1

        recommendations = [
            "• Per TOW in perdita: Aumentare % Junior o ridurre allocazione Senior",
            "• Per TOW ad alto margine: Considerare aumento peso per massimizzare profitto",
            "• Per alta concentrazione: Diversificare risorse su più TOW per ridurre rischio",
            "• Verificare che TOW complessi abbiano adeguata copertura Senior",
            "• Bilanciare mix Senior/Junior per ottimizzare costo medio ponderato"
        ]
        for rec in recommendations:
            ws.cell(row=row, column=1, value=rec)
            ws.cell(row=row, column=1).font = Font(italic=True, color='666666')
            row += 1

    # ========== SHEET 9: CALCOLO COSTI (ALL FORMULAS) ==========
    def _create_cost_calc_sheet(self):
        ws = self.wb.create_sheet("CALCOLO_COSTI")
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50

        ws['A1'] = "CALCOLO COSTI"
        ws['A1'].font = TITLE_FONT
        ws['A2'] = "TUTTI i valori in questa tabella sono FORMULE che si ricalcolano automaticamente."
        ws['A2'].font = Font(italic=True, size=10, color='008000')

        row = 4
        ws['A' + str(row)] = "COSTI DIRETTI"
        ws['A' + str(row)].font = SECTION_FONT
        row += 1

        # Costo Team (from MAPPING sheet now)
        ws['A' + str(row)] = "Costo Team"
        ws['B' + str(row)] = f"={self.named_ranges.get('TEAM_COST', '0')}"
        ws['B' + str(row)].number_format = '#,##0'
        self._style_link_cell(ws['B' + str(row)])
        ws['C' + str(row)] = "Formula: Link a MAPPING!Totale Costo"
        ws['C' + str(row)].font = Font(italic=True, color='666666')
        team_cost_row = row
        row += 2

        ws['A' + str(row)] = "OVERHEAD"
        ws['A' + str(row)].font = SECTION_FONT
        row += 1

        # Governance
        ws['A' + str(row)] = "Governance"
        ws['B' + str(row)] = f"=B{team_cost_row}*{self.named_ranges['GOVERNANCE_PCT']}"
        ws['B' + str(row)].number_format = '#,##0'
        self._style_formula_cell(ws['B' + str(row)])
        ws['C' + str(row)] = f"Formula: Team Cost × Governance%"
        ws['C' + str(row)].font = Font(italic=True, color='666666')
        gov_row = row
        row += 1

        # Risk
        ws['A' + str(row)] = "Risk Contingency"
        ws['B' + str(row)] = f"=(B{team_cost_row}+B{gov_row})*{self.named_ranges['RISK_PCT']}"
        ws['B' + str(row)].number_format = '#,##0'
        self._style_formula_cell(ws['B' + str(row)])
        ws['C' + str(row)] = f"Formula: (Team + Gov) × Risk%"
        ws['C' + str(row)].font = Font(italic=True, color='666666')
        risk_row = row
        row += 1

        # Subappalto
        ws['A' + str(row)] = "Subappalto"
        ws['B' + str(row)] = f"={self.named_ranges.get('SUB_COST', '0')}"
        ws['B' + str(row)].number_format = '#,##0'
        self._style_link_cell(ws['B' + str(row)])
        ws['C' + str(row)] = "Formula: Link a SUBAPPALTO!Totale"
        ws['C' + str(row)].font = Font(italic=True, color='666666')
        sub_row = row
        row += 2

        # TOTALE COSTI
        ws['A' + str(row)] = "TOTALE COSTI"
        ws['A' + str(row)].font = BOLD_FONT
        ws['B' + str(row)] = f"=B{team_cost_row}+B{gov_row}+B{risk_row}+B{sub_row}"
        ws['B' + str(row)].number_format = '#,##0'
        ws['B' + str(row)].font = BOLD_FONT
        ws['B' + str(row)].fill = LIGHT_FILL
        ws['B' + str(row)].border = THIN_BORDER
        ws['C' + str(row)] = "Formula: Team + Gov + Risk + Sub"
        ws['C' + str(row)].font = Font(italic=True, color='666666')
        self.named_ranges['TOTAL_COST'] = f"CALCOLO_COSTI!$B${row}"

    # ========== SHEET 8: CONTO ECONOMICO (P&L) ==========
    def _create_pl_sheet(self):
        ws = self.wb.create_sheet("CONTO_ECONOMICO")
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50

        ws['A1'] = "CONTO ECONOMICO DI COMMESSA"
        ws['A1'].font = TITLE_FONT
        ws['A2'] = "Tutti i calcoli sono formule. Modifica i PARAMETRI per vedere i risultati."
        ws['A2'].font = Font(italic=True, size=10, color='008000')

        row = 4

        # RICAVI
        ws['A' + str(row)] = "RICAVI"
        ws['A' + str(row)].font = SECTION_FONT
        row += 1

        ws['A' + str(row)] = "Base d'asta"
        ws['B' + str(row)] = f"={self.named_ranges['BASE_ASTA']}"
        ws['B' + str(row)].number_format = '#,##0'
        self._style_link_cell(ws['B' + str(row)])
        row += 1

        ws['A' + str(row)] = "Quota Lutech (se RTI)"
        ws['B' + str(row)] = f"=IF({self.named_ranges['RTI_ATTIVO']}=1,{self.named_ranges['QUOTA_LUTECH']},1)"
        ws['B' + str(row)].number_format = '0.0%'
        self._style_formula_cell(ws['B' + str(row)])
        row += 1

        ws['A' + str(row)] = "Base Effettiva"
        ws['B' + str(row)] = f"={self.named_ranges['BASE_EFFETTIVA']}"
        ws['B' + str(row)].number_format = '#,##0'
        self._style_link_cell(ws['B' + str(row)])
        row += 1

        ws['A' + str(row)] = "Sconto"
        ws['B' + str(row)] = f"={self.named_ranges['SCONTO']}"
        ws['B' + str(row)].number_format = '0.0%'
        self._style_link_cell(ws['B' + str(row)])
        row += 1

        ws['A' + str(row)] = "REVENUE"
        ws['A' + str(row)].font = BOLD_FONT
        ws['B' + str(row)] = f"={self.named_ranges['REVENUE']}"
        ws['B' + str(row)].number_format = '#,##0'
        ws['B' + str(row)].font = BOLD_FONT
        ws['B' + str(row)].fill = LIGHT_FILL
        self._style_link_cell(ws['B' + str(row)])
        revenue_row = row
        row += 2

        # COSTI
        ws['A' + str(row)] = "COSTI"
        ws['A' + str(row)].font = SECTION_FONT
        row += 1

        ws['A' + str(row)] = "Totale Costi"
        ws['B' + str(row)] = f"={self.named_ranges['TOTAL_COST']}"
        ws['B' + str(row)].number_format = '#,##0'
        self._style_link_cell(ws['B' + str(row)])
        cost_row = row
        row += 2

        # MARGINE
        ws['A' + str(row)] = "MARGINE"
        ws['A' + str(row)].font = SECTION_FONT
        row += 1

        ws['A' + str(row)] = "Margine (€)"
        ws['B' + str(row)] = f"=B{revenue_row}-B{cost_row}"
        ws['B' + str(row)].number_format = '#,##0'
        ws['B' + str(row)].font = BOLD_FONT
        self._style_formula_cell(ws['B' + str(row)])
        margin_row = row
        row += 1

        ws['A' + str(row)] = "Margine %"
        ws['B' + str(row)] = f"=IFERROR(B{margin_row}/B{revenue_row},0)"
        ws['B' + str(row)].number_format = '0.0%'
        ws['B' + str(row)].font = BOLD_FONT
        ws['B' + str(row)].fill = LIGHT_FILL
        self._style_formula_cell(ws['B' + str(row)])
        margin_pct_row = row
        self.named_ranges['MARGIN_PCT'] = f"CONTO_ECONOMICO!$B${row}"

        # Conditional formatting
        red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        yellow_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')

        ws.conditional_formatting.add(f'B{margin_pct_row}',
            FormulaRule(formula=[f'$B${margin_pct_row}<0.1'], fill=red_fill))
        ws.conditional_formatting.add(f'B{margin_pct_row}',
            FormulaRule(formula=[f'AND($B${margin_pct_row}>=0.1,$B${margin_pct_row}<0.15)'], fill=yellow_fill))
        ws.conditional_formatting.add(f'B{margin_pct_row}',
            FormulaRule(formula=[f'$B${margin_pct_row}>=0.15'], fill=green_fill))

        row += 2

        # CALCOLATORE SCONTO
        ws['A' + str(row)] = "CALCOLATORE SCONTO PER MARGINE TARGET"
        ws['A' + str(row)].font = SECTION_FONT
        row += 2

        ws['A' + str(row)] = "Margine Target"
        ws['B' + str(row)] = f"={self.named_ranges['MARGINE_TARGET']}"
        ws['B' + str(row)].number_format = '0.0%'
        self._style_link_cell(ws['B' + str(row)])
        target_row = row
        row += 1

        # Sconto necessario: discount = 1 - cost / (base * (1 - target))
        ws['A' + str(row)] = "Sconto Necessario"
        ws['B' + str(row)] = f"=MAX(0,1-{self.named_ranges['TOTAL_COST']}/({self.named_ranges['BASE_EFFETTIVA']}*(1-B{target_row})))"
        ws['B' + str(row)].number_format = '0.00%'
        ws['B' + str(row)].font = BOLD_FONT
        ws['B' + str(row)].fill = LIGHT_FILL
        self._style_formula_cell(ws['B' + str(row)])
        ws['C' + str(row)] = "Formula: 1 - Costi / (Base × (1 - Target))"
        ws['C' + str(row)].font = Font(italic=True, color='666666')

    # ========== SHEET 9: SCENARI ==========
    def _create_scenarios_sheet(self):
        ws = self.wb.create_sheet("SCENARI")
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        ws['A1'] = "ANALISI SCENARI"
        ws['A1'].font = TITLE_FONT

        row = 3

        if not self.scenarios:
            ws['A' + str(row)] = "Nessuno scenario disponibile"
            return

        headers = ['Scenario', 'Reuse %', 'Vol. Adj', 'Costo Tot (€)', 'Margine %', 'Sconto Sugg.']
        self._add_header_row(ws, row, headers)
        row += 1
        data_start = row

        for s in self.scenarios:
            ws.cell(row=row, column=1, value=s.get('name', '')).border = THIN_BORDER

            reuse = s.get('reuse_factor', 0)
            ws.cell(row=row, column=2, value=reuse if reuse <= 1 else reuse/100)
            ws.cell(row=row, column=2).number_format = '0.0%'
            ws.cell(row=row, column=2).border = THIN_BORDER

            vol = s.get('volume_adjustment', 1)
            ws.cell(row=row, column=3, value=1 - vol if vol < 1 else 0)
            ws.cell(row=row, column=3).number_format = '0.0%'
            ws.cell(row=row, column=3).border = THIN_BORDER

            ws.cell(row=row, column=4, value=s.get('total_cost', 0))
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=4).border = THIN_BORDER

            margin = s.get('margin_pct', 0)
            ws.cell(row=row, column=5, value=margin/100 if margin > 1 else margin)
            ws.cell(row=row, column=5).number_format = '0.0%'
            ws.cell(row=row, column=5).border = THIN_BORDER

            disc = s.get('suggested_discount', 0)
            ws.cell(row=row, column=6, value=disc/100 if disc > 1 else disc)
            ws.cell(row=row, column=6).number_format = '0.0%'
            ws.cell(row=row, column=6).border = THIN_BORDER

            row += 1

        # Color scale for margin
        color_scale = ColorScaleRule(
            start_type='num', start_value=0, start_color='DC3545',
            mid_type='num', mid_value=0.15, mid_color='FFC107',
            end_type='num', end_value=0.30, end_color='28A745'
        )
        ws.conditional_formatting.add(f'E{data_start}:E{row-1}', color_scale)

    # ========== SHEET 10: SCHEMA OFFERTA (PxQ) ==========
    def _create_offer_sheet(self):
        ws = self.wb.create_sheet("SCHEMA_OFFERTA")
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 38
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 14
        ws.freeze_panes = 'A4'

        ws['A1'] = "SCHEMA OFFERTA ECONOMICA"
        ws['A1'].font = TITLE_FONT
        ws['A2'] = "Prezzi calcolati con FORMULE. Modifica Quota % per ribilanciare."
        ws['A2'].font = Font(italic=True, size=10, color='008000')

        row = 4
        headers = ['TOW ID', 'Descrizione', 'Tipo', 'Quantità', 'Prezzo Unit.', 'Prezzo Tot.', 'Quota %']
        self._add_header_row(ws, row, headers)
        row += 1
        data_start = row

        tows = self.bp.get('tows') or self.bp.get('tow_config') or []

        # Calculate initial shares from breakdown
        # Handle both formats: {TOW_01: 350000} and {TOW_01: {'cost': 350000}}
        def get_tow_cost(tid):
            val = self.tow_breakdown.get(tid, 0)
            if isinstance(val, dict):
                return val.get('cost', 0)
            return val or 0

        total_breakdown_cost = sum(
            get_tow_cost(tid) for tid in self.tow_breakdown.keys() if tid != '__no_tow__'
        ) or 1

        for tow in tows:
            tow_id = tow.get('tow_id', tow.get('id', ''))
            tow_cost = get_tow_cost(tow_id)
            share = tow_cost / total_breakdown_cost

            ws.cell(row=row, column=1, value=tow_id).border = THIN_BORDER
            ws.cell(row=row, column=2, value=tow.get('label', '')).border = THIN_BORDER

            cell_type = ws.cell(row=row, column=3, value=tow.get('type', 'task'))
            cell_type.alignment = CENTER
            cell_type.border = THIN_BORDER

            # Quantity based on type
            if tow.get('type') == 'task':
                qty = tow.get('num_tasks', 1) or 1
            elif tow.get('type') == 'corpo':
                qty = self.bp.get('duration_months', 36)
            else:
                qty = 1

            cell_qty = ws.cell(row=row, column=4, value=qty)
            self._style_input_cell(cell_qty)
            cell_qty.alignment = CENTER

            # Quota % (input)
            cell_quota = ws.cell(row=row, column=7, value=share)
            cell_quota.number_format = '0.0%'
            self._style_input_cell(cell_quota)
            cell_quota.alignment = CENTER

            # Prezzo Tot = Revenue × Quota (FORMULA)
            cell_tot = ws.cell(row=row, column=6)
            cell_tot.value = f"={self.named_ranges['REVENUE']}*G{row}"
            cell_tot.number_format = '#,##0'
            self._style_formula_cell(cell_tot)

            # Prezzo Unit = Tot / Qty (FORMULA)
            cell_unit = ws.cell(row=row, column=5)
            cell_unit.value = f"=IFERROR(F{row}/D{row},0)"
            cell_unit.number_format = '#,##0.00'
            self._style_formula_cell(cell_unit)

            row += 1

        # Totals
        if tows:
            ws.cell(row=row, column=1, value="TOTALE").font = BOLD_FONT
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row=row, column=1).border = THIN_BORDER

            ws.cell(row=row, column=6, value=f"=SUM(F{data_start}:F{row-1})")
            ws.cell(row=row, column=6).number_format = '#,##0'
            ws.cell(row=row, column=6).font = BOLD_FONT
            ws.cell(row=row, column=6).fill = LIGHT_FILL
            ws.cell(row=row, column=6).border = THIN_BORDER

            ws.cell(row=row, column=7, value=f"=SUM(G{data_start}:G{row-1})")
            ws.cell(row=row, column=7).number_format = '0.0%'
            ws.cell(row=row, column=7).font = BOLD_FONT
            ws.cell(row=row, column=7).border = THIN_BORDER

            row += 2

            # Validation check
            ws.cell(row=row, column=1, value="VERIFICA: La somma quote deve essere 100%")
            ws.cell(row=row, column=1).font = Font(italic=True, color='666666')
            ws.cell(row=row, column=7, value=f"=IF(G{row-2}=1,\"OK\",\"ERRORE!\")")
            ws.cell(row=row, column=7).font = BOLD_FONT


def generate_business_plan_excel(
    lot_key: str,
    business_plan: Dict[str, Any],
    costs: Dict[str, float],
    clean_team_cost: float,
    base_amount: float,
    is_rti: bool = False,
    quota_lutech: float = 1.0,
    scenarios: Optional[List[Dict[str, Any]]] = None,
    tow_breakdown: Optional[Dict[str, Any]] = None,
    profile_rates: Optional[Dict[str, float]] = None,
    intervals: Optional[List[Dict[str, Any]]] = None,
    lutech_breakdown: Optional[Dict[str, Any]] = None,
) -> io.BytesIO:
    generator = BusinessPlanExcelGenerator(
        lot_key=lot_key,
        business_plan=business_plan,
        costs=costs,
        clean_team_cost=clean_team_cost,
        base_amount=base_amount,
        is_rti=is_rti,
        quota_lutech=quota_lutech,
        scenarios=scenarios,
        tow_breakdown=tow_breakdown,
        profile_rates=profile_rates,
        intervals=intervals,
        lutech_breakdown=lutech_breakdown,
    )
    return generator.generate()
