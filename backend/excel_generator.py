"""
Excel Report Generator for Simulator Poste
Generates professional multi-sheet Excel reports with formulas and advanced formatting
"""

import io
from datetime import datetime
from typing import Dict, Any, List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle, GradientFill
)
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# ============================================================================
# STYLE DEFINITIONS
# ============================================================================

# Colors
COLORS = {
    'primary': '1E3A5F',
    'primary_light': '2563EB',
    'secondary': 'FFCC00',
    'success': '28A745',
    'warning': 'FFC107',
    'danger': 'DC3545',
    'light': 'F8F9FA',
    'dark': '333333',
    'muted': '6C757D',
    'white': 'FFFFFF',
    'input_bg': 'FFF9C4',
    'formula_bg': 'E3F2FD',
    'lutech': '0066CC',
    'partner_1': '6366F1',
    'partner_2': '8B5CF6',
    'partner_3': 'EC4899',
}

# Borders
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

MEDIUM_BORDER = Border(
    left=Side(style='medium', color='333333'),
    right=Side(style='medium', color='333333'),
    top=Side(style='medium', color='333333'),
    bottom=Side(style='medium', color='333333')
)

# Fills
HEADER_FILL = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
LIGHT_FILL = PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid')
INPUT_FILL = PatternFill(start_color=COLORS['input_bg'], end_color=COLORS['input_bg'], fill_type='solid')
FORMULA_FILL = PatternFill(start_color=COLORS['formula_bg'], end_color=COLORS['formula_bg'], fill_type='solid')

# Fonts
TITLE_FONT = Font(name='Calibri', size=20, bold=True, color=COLORS['primary'])
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color=COLORS['white'])
SECTION_FONT = Font(name='Calibri', size=14, bold=True, color=COLORS['primary'])
LABEL_FONT = Font(name='Calibri', size=10, bold=True, color=COLORS['dark'])
VALUE_FONT = Font(name='Calibri', size=11, color=COLORS['dark'])
FORMULA_FONT = Font(name='Consolas', size=10, color=COLORS['muted'])
KPI_FONT = Font(name='Calibri', size=24, bold=True, color=COLORS['primary'])
KPI_LABEL_FONT = Font(name='Calibri', size=9, color=COLORS['muted'])

# Alignments
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)


class ExcelReportGenerator:
    """Generates professional Excel reports with formulas and charts"""

    def __init__(
        self,
        lot_key: str,
        lot_config: Dict[str, Any],
        base_amount: float,
        my_discount: float,
        competitor_discount: float,
        technical_score: float,
        economic_score: float,
        total_score: float,
        details: Dict[str, float],
        weighted_scores: Dict[str, float],
        category_scores: Dict[str, float],
        max_tech_score: float,
        max_econ_score: float,
        alpha: float,
        win_probability: float,
        tech_inputs_full: Optional[Dict[str, Any]] = None,
        rti_quotas: Optional[Dict[str, float]] = None,
    ):
        self.lot_key = lot_key
        self.lot_config = lot_config
        self.base_amount = base_amount
        self.my_discount = my_discount
        self.competitor_discount = competitor_discount
        self.technical_score = technical_score
        self.economic_score = economic_score
        self.total_score = total_score
        self.details = details
        self.weighted_scores = weighted_scores
        self.category_scores = category_scores
        self.max_tech_score = max_tech_score
        self.max_econ_score = max_econ_score
        self.alpha = alpha
        self.win_probability = win_probability
        self.tech_inputs_full = tech_inputs_full or {}
        self.rti_quotas = rti_quotas or {}
        
        self.is_rti = lot_config.get('rti_enabled', False)
        self.rti_companies = ['Lutech'] + (lot_config.get('rti_companies', []) or [])
        
        self.wb = Workbook()
        self.named_ranges = {}
        
        # Company colors for RTI
        self.company_colors = {'Lutech': COLORS['lutech']}
        partner_colors = [COLORS['partner_1'], COLORS['partner_2'], COLORS['partner_3']]
        for i, company in enumerate(self.lot_config.get('rti_companies', []) or []):
            self.company_colors[company] = partner_colors[i % len(partner_colors)]

    def generate(self) -> io.BytesIO:
        """Generate the complete Excel report"""
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Create sheets in order (we need Tecnico and Economico first for Dashboard refs)
        self._create_technical_sheet()
        self._create_economic_sheet()
        
        if self.is_rti:
            self._create_rti_sheet()
        
        self._create_dashboard_sheet()
        self._create_config_sheet()
        self._create_named_ranges()
        
        # Hide gridlines on all sheets
        for sheet_name in self.wb.sheetnames:
            self.wb[sheet_name].sheet_view.showGridLines = False
        
        self.wb.active = self.wb['Dashboard']
        
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _create_dashboard_sheet(self):
        """Create the executive dashboard sheet"""
        ws = self.wb.create_sheet('Dashboard', 0)  # Insert at position 0
        ws.sheet_properties.tabColor = COLORS['primary']
        
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 3
        
        row = 2
        
        # Header
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = f'REPORT - {self.lot_key}'
        ws[f'B{row}'].font = TITLE_FONT
        ws[f'B{row}'].alignment = CENTER
        row += 1
        
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = f'Generato il {datetime.now().strftime("%d/%m/%Y alle %H:%M")}'
        ws[f'B{row}'].font = Font(size=10, italic=True, color=COLORS['muted'])
        ws[f'B{row}'].alignment = CENTER
        row += 3
        
        # KPI boxes - only 4: Total, Tecnico, Economico, Sconto (no Win%)
        kpi_row = row
        
        # PUNTEGGIO TOTALE - formula referencing Tecnico + Economico
        ws[f'B{kpi_row}'] = 'PUNTEGGIO TOTALE'
        ws[f'B{kpi_row}'].font = KPI_LABEL_FONT
        ws[f'B{kpi_row}'].alignment = CENTER
        ws[f'B{kpi_row}'].border = THIN_BORDER
        # Formula: Tecnico total + Economico score
        ws[f'B{kpi_row+1}'] = f"=Tecnico!G{self.tech_cat_total_row}+Economico!C{self.econ_score_row}"
        ws[f'B{kpi_row+1}'].font = KPI_FONT
        ws[f'B{kpi_row+1}'].alignment = CENTER
        ws[f'B{kpi_row+1}'].number_format = '0.00'
        ws[f'B{kpi_row+1}'].fill = FORMULA_FILL
        ws[f'B{kpi_row+1}'].border = THIN_BORDER
        
        # TECNICO - formula referencing Tecnico sheet
        ws[f'C{kpi_row}'] = 'TECNICO'
        ws[f'C{kpi_row}'].font = KPI_LABEL_FONT
        ws[f'C{kpi_row}'].alignment = CENTER
        ws[f'C{kpi_row}'].border = THIN_BORDER
        ws[f'C{kpi_row+1}'] = f"=Tecnico!G{self.tech_cat_total_row}"
        ws[f'C{kpi_row+1}'].font = Font(size=20, bold=True, color=COLORS['primary_light'])
        ws[f'C{kpi_row+1}'].alignment = CENTER
        ws[f'C{kpi_row+1}'].number_format = '0.00'
        ws[f'C{kpi_row+1}'].fill = FORMULA_FILL
        ws[f'C{kpi_row+1}'].border = THIN_BORDER
        
        # ECONOMICO - formula referencing Economico sheet
        ws[f'D{kpi_row}'] = 'ECONOMICO'
        ws[f'D{kpi_row}'].font = KPI_LABEL_FONT
        ws[f'D{kpi_row}'].alignment = CENTER
        ws[f'D{kpi_row}'].border = THIN_BORDER
        ws[f'D{kpi_row+1}'] = f"=Economico!C{self.econ_score_row}"
        ws[f'D{kpi_row+1}'].font = Font(size=20, bold=True, color=COLORS['secondary'])
        ws[f'D{kpi_row+1}'].alignment = CENTER
        ws[f'D{kpi_row+1}'].number_format = '0.00'
        ws[f'D{kpi_row+1}'].fill = FORMULA_FILL
        ws[f'D{kpi_row+1}'].border = THIN_BORDER
        
        # SCONTO - formula referencing Economico sheet
        ws[f'E{kpi_row}'] = 'SCONTO'
        ws[f'E{kpi_row}'].font = KPI_LABEL_FONT
        ws[f'E{kpi_row}'].alignment = CENTER
        ws[f'E{kpi_row}'].border = THIN_BORDER
        ws[f'E{kpi_row+1}'] = f"=Economico!C{self.econ_sconto_row}/100"
        ws[f'E{kpi_row+1}'].font = Font(size=20, bold=True, color=COLORS['dark'])
        ws[f'E{kpi_row+1}'].alignment = CENTER
        ws[f'E{kpi_row+1}'].number_format = '0.0%'
        ws[f'E{kpi_row+1}'].fill = FORMULA_FILL
        ws[f'E{kpi_row+1}'].border = THIN_BORDER
        
        row = kpi_row + 4
        
        # Input section (simplified - main inputs are in Economico)
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = 'PARAMETRI DI INPUT'
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        # Reference to Economico sheet values
        ws[f'B{row}'] = 'Base d\'Asta'
        ws[f'B{row}'].font = LABEL_FONT
        ws[f'C{row}'] = f"=Economico!C{self.econ_base_row}"
        ws[f'C{row}'].fill = FORMULA_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].number_format = '€ #,##0.00'
        row += 1
        
        ws[f'B{row}'] = 'Sconto (%)'
        ws[f'B{row}'].font = LABEL_FONT
        ws[f'C{row}'] = f"=Economico!C{self.econ_sconto_row}"
        ws[f'C{row}'].fill = FORMULA_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].number_format = '0.0'
        row += 1
        
        ws[f'B{row}'] = 'Sconto Best Offer (%)'
        ws[f'B{row}'].font = LABEL_FONT
        ws[f'C{row}'] = f"=Economico!C{self.econ_sconto_best_row}"
        ws[f'C{row}'].fill = FORMULA_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].number_format = '0.0'
        row += 1
        
        ws[f'B{row}'] = 'Alpha (α)'
        ws[f'B{row}'].font = LABEL_FONT
        ws[f'C{row}'] = f"=Economico!C{self.econ_alpha_row}"
        ws[f'C{row}'].fill = FORMULA_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].number_format = '0.00'
        row += 2
        
        ws.freeze_panes = 'B5'

    def _create_technical_sheet(self):
        """Create the technical score analysis sheet with formulas"""
        ws = self.wb.create_sheet('Tecnico')
        ws.sheet_properties.tabColor = COLORS['primary_light']
        
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 14
        ws.column_dimensions['L'].width = 8
        
        row = 2
        
        ws.merge_cells(f'B{row}:L{row}')
        ws[f'B{row}'] = 'ANALISI PUNTEGGIO TECNICO'
        ws[f'B{row}'].font = TITLE_FONT
        row += 2
        
        ws[f'B{row}'] = 'Punteggio Tecnico Totale:'
        ws[f'B{row}'].font = LABEL_FONT
        tech_total_row = row
        tech_total_col = 'C'
        ws[f'{tech_total_col}{row}'] = self.technical_score  # Will be updated to formula later
        ws[f'{tech_total_col}{row}'].font = Font(size=16, bold=True, color=COLORS['primary'])
        ws[f'{tech_total_col}{row}'].number_format = '0.00'
        
        ws[f'E{row}'] = 'Max Ottenibile:'
        ws[f'E{row}'].font = LABEL_FONT
        ws[f'F{row}'] = self.max_tech_score
        ws[f'F{row}'].number_format = '0.00'
        self.named_ranges['MaxTech'] = f"'Tecnico'!$F${row}"
        max_tech_row = row
        
        ws[f'H{row}'] = 'Raggiungimento:'
        ws[f'H{row}'].font = LABEL_FONT
        pct = self.technical_score / self.max_tech_score if self.max_tech_score > 0 else 0
        ws[f'I{row}'] = f'=C{row}/F{row}'
        ws[f'I{row}'].number_format = '0.0%'
        ws[f'I{row}'].fill = FORMULA_FILL
        if pct >= 0.7:
            ws[f'I{row}'].font = Font(bold=True, color=COLORS['success'])
        elif pct >= 0.5:
            ws[f'I{row}'].font = Font(bold=True, color=COLORS['warning'])
        else:
            ws[f'I{row}'].font = Font(bold=True, color=COLORS['danger'])
        row += 3
        
        # Category breakdown with formulas
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = 'BREAKDOWN PER CATEGORIA'
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        cat_headers = ['Categoria', 'Punteggio Raw', 'Max Raw', '%', 'Peso Gara', 'Punteggio Pesato (Formula)']
        for col, header in enumerate(cat_headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        cat_header_row = row
        row += 1
        
        # Track category data rows for formulas
        cat_rows = {}
        reqs = self.lot_config.get('reqs', [])
        
        # Calculate category aggregates
        category_data = {
            'company_certs': {'raw': 0, 'max_raw': 0, 'gara_weight': 0},
            'resource': {'raw': 0, 'max_raw': 0, 'gara_weight': 0},
            'reference': {'raw': 0, 'max_raw': 0, 'gara_weight': 0},
            'project': {'raw': 0, 'max_raw': 0, 'gara_weight': 0},
        }
        
        company_certs = self.lot_config.get('company_certs', [])
        # For company_certs, raw is the SUM of individual cert points obtained
        # max_raw is the SUM of individual max points, gara_weight is total weight
        company_certs_raw_sum = 0
        for cert in company_certs:
            cert_id = cert.get('id', '')
            # Check if company has this cert (stored in tech_inputs_full)
            cert_input = self.tech_inputs_full.get(f'company_cert_{cert_id}', {})
            has_cert = cert_input.get('has_cert', False) if isinstance(cert_input, dict) else cert_input
            if has_cert:
                company_certs_raw_sum += cert.get('points', 0)
            category_data['company_certs']['max_raw'] += cert.get('points', 0)
            category_data['company_certs']['gara_weight'] += cert.get('gara_weight', 0)
        category_data['company_certs']['raw'] = company_certs_raw_sum
        
        for req in reqs:
            req_type = req.get('type', 'resource')
            if req_type in category_data:
                category_data[req_type]['raw'] += self.details.get(req.get('id', ''), 0)
                category_data[req_type]['max_raw'] += req.get('max_points', 0)
                category_data[req_type]['gara_weight'] += req.get('gara_weight', 0)
        
        cat_labels = {
            'company_certs': 'Certificazioni Aziendali',
            'resource': 'Certificazioni Professionali',
            'reference': 'Referenze',
            'project': 'Progetti Tecnici'
        }
        
        cat_start_row = row
        for cat_key, cat_label in cat_labels.items():
            data = category_data[cat_key]
            cat_rows[cat_key] = row
            
            ws.cell(row=row, column=2, value=cat_label).font = LABEL_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            # Raw score
            raw_cell = ws.cell(row=row, column=3, value=data['raw'])
            raw_cell.number_format = '0.00'
            raw_cell.fill = INPUT_FILL
            raw_cell.border = THIN_BORDER
            # Max raw
            max_raw_cell = ws.cell(row=row, column=4, value=data['max_raw'])
            max_raw_cell.number_format = '0.00'
            max_raw_cell.border = THIN_BORDER
            # Percentage - formula
            pct_cell = ws.cell(row=row, column=5, value=f'=IF(D{row}=0,0,C{row}/D{row})')
            pct_cell.number_format = '0.0%'
            pct_cell.fill = FORMULA_FILL
            pct_cell.border = THIN_BORDER
            # Gara weight
            weight_cell = ws.cell(row=row, column=6, value=data['gara_weight'])
            weight_cell.number_format = '0.00'
            weight_cell.border = THIN_BORDER
            # Weighted score - FORMULA for ALL categories: raw * gara_weight / max_raw
            weighted_cell = ws.cell(row=row, column=7, value=f'=IF(D{row}=0,0,C{row}*F{row}/D{row})')
            weighted_cell.number_format = '0.00'
            weighted_cell.fill = FORMULA_FILL
            weighted_cell.font = Font(bold=True, color=COLORS['primary'])
            weighted_cell.border = THIN_BORDER
            row += 1
        
        # Total row with formulas
        ws.cell(row=row, column=2, value='TOTALE').font = Font(bold=True)
        ws.cell(row=row, column=2).border = MEDIUM_BORDER
        ws.cell(row=row, column=2).fill = LIGHT_FILL
        ws.cell(row=row, column=3, value=f'=SUM(C{cat_start_row}:C{row-1})').number_format = '0.00'
        ws.cell(row=row, column=3).border = MEDIUM_BORDER
        ws.cell(row=row, column=3).fill = LIGHT_FILL
        ws.cell(row=row, column=4, value=f'=SUM(D{cat_start_row}:D{row-1})').number_format = '0.00'
        ws.cell(row=row, column=4).border = MEDIUM_BORDER
        ws.cell(row=row, column=4).fill = LIGHT_FILL
        ws.cell(row=row, column=5, value=f'=IF(D{row}=0,0,C{row}/D{row})').number_format = '0.0%'
        ws.cell(row=row, column=5).fill = FORMULA_FILL
        ws.cell(row=row, column=5).border = MEDIUM_BORDER
        ws.cell(row=row, column=6, value=f'=SUM(F{cat_start_row}:F{row-1})').number_format = '0.00'
        ws.cell(row=row, column=6).border = MEDIUM_BORDER
        ws.cell(row=row, column=6).fill = LIGHT_FILL
        ws.cell(row=row, column=7, value=f'=SUM(G{cat_start_row}:G{row-1})').number_format = '0.00'
        ws.cell(row=row, column=7).font = Font(bold=True, color=COLORS['primary'])
        ws.cell(row=row, column=7).fill = FORMULA_FILL
        ws.cell(row=row, column=7).border = MEDIUM_BORDER
        
        # Update tech total to reference this sum
        ws[f'{tech_total_col}{tech_total_row}'] = f'=G{row}'
        ws[f'{tech_total_col}{tech_total_row}'].fill = FORMULA_FILL
        
        cat_total_row = row
        # Store for Dashboard reference
        self.tech_cat_total_row = cat_total_row
        row += 3
        
        # Detailed requirements with expanded certifications
        ws.merge_cells(f'B{row}:L{row}')
        ws[f'B{row}'] = 'DETTAGLIO REQUISITI E CERTIFICAZIONI'
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        det_headers = ['ID', 'Requisito', 'Certificazione', 'Azienda', 'Tipo', 'Score Raw', 'Max Raw', '%', 'Peso Gara', 'Score Pesato', 'Status']
        for col, header in enumerate(det_headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        row += 1
        
        req_start_row = row
        type_labels = {'resource': 'Cert. Prof.', 'reference': 'Referenza', 'project': 'Progetto'}
        
        for req in reqs:
            req_id = req.get('id', '')
            req_type = req.get('type', 'resource')
            raw_score = self.details.get(req_id, 0)
            max_score = req.get('max_points', 0)
            gara_weight = req.get('gara_weight', 0)
            tech_input = self.tech_inputs_full.get(req_id, {})
            
            if req_type == 'resource':
                # Expand by certification and company
                cert_company_counts = tech_input.get('cert_company_counts', {})
                selected_prof_certs = req.get('selected_prof_certs', [])
                
                # Get all certifications with their company assignments
                cert_entries = []
                for cert_name, company_counts in cert_company_counts.items():
                    if isinstance(company_counts, dict):
                        for company, count in company_counts.items():
                            if count > 0:
                                cert_entries.append({
                                    'cert_name': cert_name,
                                    'company': company,
                                    'count': count
                                })
                
                # If no cert entries but we have selected_prof_certs, use them
                if not cert_entries and selected_prof_certs:
                    for cert in selected_prof_certs:
                        cert_entries.append({
                            'cert_name': cert,
                            'company': 'Lutech',
                            'count': 0
                        })
                
                # If still no entries, create a single row
                if not cert_entries:
                    cert_entries = [{'cert_name': '-', 'company': 'Lutech', 'count': 0}]
                
                # Write a row for each certification
                first_row_for_req = row
                for i, entry in enumerate(cert_entries):
                    is_first = (i == 0)
                    
                    # ID (only first row)
                    id_cell = ws.cell(row=row, column=2, value=req_id if is_first else '')
                    id_cell.font = Font(size=9, color=COLORS['muted'])
                    
                    # Requisito (only first row)
                    ws.cell(row=row, column=3, value=req.get('label', '') if is_first else '').alignment = WRAP
                    
                    # Certificazione
                    ws.cell(row=row, column=4, value=entry['cert_name'])
                    
                    # Azienda
                    company = entry['company']
                    company_cell = ws.cell(row=row, column=5, value=company)
                    company_cell.font = Font(bold=True, color=self.company_colors.get(company, COLORS['dark']))
                    
                    # Tipo
                    ws.cell(row=row, column=6, value=type_labels.get(req_type, req_type) if is_first else '')
                    
                    # Score Raw (only first row, with formula reference)
                    if is_first:
                        ws.cell(row=row, column=7, value=raw_score).number_format = '0.00'
                        ws.cell(row=row, column=7).fill = INPUT_FILL
                    
                    # Max Raw (only first row)
                    if is_first:
                        ws.cell(row=row, column=8, value=max_score).number_format = '0.00'
                    
                    # % (only first row - formula)
                    if is_first:
                        ws.cell(row=row, column=9, value=f'=IF(H{row}=0,0,G{row}/H{row})').number_format = '0.0%'
                        ws.cell(row=row, column=9).fill = FORMULA_FILL
                    
                    # Peso Gara (only first row)
                    if is_first:
                        ws.cell(row=row, column=10, value=gara_weight).number_format = '0.00'
                    
                    # Score Pesato - FORMULA (only first row)
                    if is_first:
                        ws.cell(row=row, column=11, value=f'=IF(H{row}=0,0,G{row}*J{row}/H{row})').number_format = '0.00'
                        ws.cell(row=row, column=11).fill = FORMULA_FILL
                        ws.cell(row=row, column=11).font = Font(bold=True, color=COLORS['primary'])
                    
                    # Status (only first row)
                    if is_first:
                        req_pct = raw_score / max_score if max_score > 0 else 0
                        if req_pct >= 0.8:
                            ws.cell(row=row, column=12, value='OK')
                        elif req_pct >= 0.5:
                            ws.cell(row=row, column=12, value='WARN')
                        elif req_pct > 0:
                            ws.cell(row=row, column=12, value='LOW')
                        else:
                            ws.cell(row=row, column=12, value='MISS')
                    
                    for col in range(2, 13):
                        ws.cell(row=row, column=col).border = THIN_BORDER
                    row += 1
                    
            else:
                # Non-resource types (reference, project) - single row
                assigned = tech_input.get('assigned_company', '') or 'Lutech'
                
                ws.cell(row=row, column=2, value=req_id).font = Font(size=9, color=COLORS['muted'])
                ws.cell(row=row, column=3, value=req.get('label', '')).alignment = WRAP
                ws.cell(row=row, column=4, value='-')  # No certification
                
                company_cell = ws.cell(row=row, column=5, value=assigned)
                company_cell.font = Font(bold=True, color=self.company_colors.get(assigned, COLORS['dark']))
                
                ws.cell(row=row, column=6, value=type_labels.get(req_type, req_type))
                ws.cell(row=row, column=7, value=raw_score).number_format = '0.00'
                ws.cell(row=row, column=7).fill = INPUT_FILL
                ws.cell(row=row, column=8, value=max_score).number_format = '0.00'
                ws.cell(row=row, column=9, value=f'=IF(H{row}=0,0,G{row}/H{row})').number_format = '0.0%'
                ws.cell(row=row, column=9).fill = FORMULA_FILL
                ws.cell(row=row, column=10, value=gara_weight).number_format = '0.00'
                # Weighted score formula
                ws.cell(row=row, column=11, value=f'=IF(H{row}=0,0,G{row}*J{row}/H{row})').number_format = '0.00'
                ws.cell(row=row, column=11).fill = FORMULA_FILL
                ws.cell(row=row, column=11).font = Font(bold=True, color=COLORS['primary'])
                
                req_pct = raw_score / max_score if max_score > 0 else 0
                if req_pct >= 0.8:
                    ws.cell(row=row, column=12, value='OK')
                elif req_pct >= 0.5:
                    ws.cell(row=row, column=12, value='WARN')
                elif req_pct > 0:
                    ws.cell(row=row, column=12, value='LOW')
                else:
                    ws.cell(row=row, column=12, value='MISS')
                
                for col in range(2, 13):
                    ws.cell(row=row, column=col).border = THIN_BORDER
                row += 1
        
        # Color scale for percentage column
        if row > req_start_row:
            rule = ColorScaleRule(
                start_type='num', start_value=0, start_color='F8D7DA',
                mid_type='num', mid_value=0.5, mid_color='FFF3CD',
                end_type='num', end_value=1, end_color='D4EDDA'
            )
            ws.conditional_formatting.add(f'I{req_start_row}:I{row-1}', rule)
        
        ws.freeze_panes = 'C5'

    def _create_economic_sheet(self):
        """Create the economic analysis sheet"""
        ws = self.wb.create_sheet('Economico')
        ws.sheet_properties.tabColor = COLORS['secondary']
        
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        
        row = 2
        
        ws.merge_cells(f'B{row}:G{row}')
        ws[f'B{row}'] = 'ANALISI ECONOMICA'
        ws[f'B{row}'].font = TITLE_FONT
        row += 2
        
        ws.merge_cells(f'B{row}:G{row}')
        ws[f'B{row}'] = 'FORMULA: Score = MaxEcon × ((BaseAsta - PrezzoMio) / (BaseAsta - PrezzoBest))^α'
        ws[f'B{row}'].font = Font(size=11, italic=True, color=COLORS['muted'])
        row += 2
        
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = 'PARAMETRI'
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        ws[f'B{row}'] = 'Base d\'Asta'
        ws[f'B{row}'].font = LABEL_FONT
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'C{row}'] = self.base_amount
        ws[f'C{row}'].fill = INPUT_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].number_format = '€ #,##0.00'
        base_row = row
        self.econ_base_row = row
        self.named_ranges['BaseAsta'] = f"'Economico'!$C${row}"
        row += 1
        
        ws[f'B{row}'] = 'Sconto (%)'
        ws[f'B{row}'].font = LABEL_FONT
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'C{row}'] = self.my_discount
        ws[f'C{row}'].fill = INPUT_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].number_format = '0.0'
        sconto_mio_row = row
        self.econ_sconto_row = row
        self.named_ranges['Sconto'] = f"'Economico'!$C${row}"
        row += 1
        
        ws[f'B{row}'] = 'Sconto Best Offer (%)'
        ws[f'B{row}'].font = LABEL_FONT
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'C{row}'] = self.competitor_discount
        ws[f'C{row}'].fill = INPUT_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].number_format = '0.0'
        sconto_best_row = row
        self.econ_sconto_best_row = row
        self.named_ranges['ScontoBest'] = f"'Economico'!$C${row}"
        row += 1
        
        ws[f'B{row}'] = 'Alpha (α)'
        ws[f'B{row}'].font = LABEL_FONT
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'C{row}'] = self.alpha
        ws[f'C{row}'].fill = INPUT_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].number_format = '0.00'
        alpha_row = row
        self.econ_alpha_row = row
        self.named_ranges['Alpha'] = f"'Economico'!$C${row}"
        row += 1
        
        ws[f'B{row}'] = 'Max Punteggio Economico'
        ws[f'B{row}'].font = LABEL_FONT
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'C{row}'] = self.max_econ_score
        ws[f'C{row}'].fill = LIGHT_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].number_format = '0.00'
        max_econ_row = row
        self.named_ranges['MaxEcon'] = f"'Economico'!$C${row}"
        row += 2
        
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = 'CALCOLI'
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        ws[f'B{row}'] = 'Prezzo Offerto'
        ws[f'B{row}'].font = LABEL_FONT
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'C{row}'] = f'=C{base_row}*(1-C{sconto_mio_row}/100)'
        ws[f'C{row}'].fill = FORMULA_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].number_format = '€ #,##0.00'
        prezzo_mio_row = row
        self.econ_prezzo_mio_row = row
        row += 1
        
        ws[f'B{row}'] = 'Prezzo Best Offer'
        ws[f'B{row}'].font = LABEL_FONT
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'C{row}'] = f'=C{base_row}*(1-C{sconto_best_row}/100)'
        ws[f'C{row}'].fill = FORMULA_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].number_format = '€ #,##0.00'
        prezzo_best_row = row
        row += 1
        
        ws[f'B{row}'] = 'Rapporto (R)'
        ws[f'B{row}'].font = LABEL_FONT
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'C{row}'] = f'=IF(C{base_row}-C{prezzo_best_row}=0,0,(C{base_row}-C{prezzo_mio_row})/(C{base_row}-C{prezzo_best_row}))'
        ws[f'C{row}'].fill = FORMULA_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].number_format = '0.0000'
        rapporto_row = row
        row += 1
        
        ws[f'B{row}'] = 'PUNTEGGIO ECONOMICO'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].border = MEDIUM_BORDER
        ws[f'C{row}'] = f'=C{max_econ_row}*(C{rapporto_row}^C{alpha_row})'
        ws[f'C{row}'].fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        ws[f'C{row}'].border = MEDIUM_BORDER
        ws[f'C{row}'].number_format = '0.00'
        ws[f'C{row}'].font = Font(bold=True, size=14, color=COLORS['success'])
        self.econ_score_row = row
        row += 3
        
        # Scenario table - discounts from 1% to 100%
        ws.merge_cells(f'B{row}:G{row}')
        ws[f'B{row}'] = 'TABELLA SCENARI (1%-100%)'
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        scenario_headers = ['Sconto %', 'Prezzo Offerto', 'Rapporto', 'Score Econ.', 'Score Tecnico', 'TOTALE']
        for col, header in enumerate(scenario_headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        row += 1
        
        scenario_start = row
        # Discounts from 1 to 100 with step 1
        for discount in range(1, 101):
            ws.cell(row=row, column=2, value=discount / 100).number_format = '0%'
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=3, value=f'=$C${base_row}*(1-B{row})').number_format = '€ #,##0.00'
            ws.cell(row=row, column=3).border = THIN_BORDER
            ws.cell(row=row, column=4, value=f'=IF($C${base_row}-$C${prezzo_best_row}=0,0,($C${base_row}-C{row})/($C${base_row}-$C${prezzo_best_row}))').number_format = '0.0000'
            ws.cell(row=row, column=4).border = THIN_BORDER
            ws.cell(row=row, column=5, value=f'=$C${max_econ_row}*(D{row}^$C${alpha_row})').number_format = '0.00'
            ws.cell(row=row, column=5).border = THIN_BORDER
            # Reference to Tecnico sheet total
            ws.cell(row=row, column=6, value=f'=Tecnico!G{self.tech_cat_total_row}').number_format = '0.00'
            ws.cell(row=row, column=6).border = THIN_BORDER
            ws.cell(row=row, column=7, value=f'=E{row}+F{row}').number_format = '0.00'
            ws.cell(row=row, column=7).font = Font(bold=True)
            ws.cell(row=row, column=7).border = THIN_BORDER
            
            # Highlight current discount row
            if discount == int(self.my_discount):
                for col in range(2, 8):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
            row += 1
        
        rule = ColorScaleRule(
            start_type='min', start_color='F8D7DA',
            mid_type='percentile', mid_value=50, mid_color='FFF3CD',
            end_type='max', end_color='D4EDDA'
        )
        ws.conditional_formatting.add(f'G{scenario_start}:G{row-1}', rule)
        
        ws.freeze_panes = 'B5'

    def _create_rti_sheet(self):
        """Create the RTI contributions breakdown sheet"""
        from openpyxl.worksheet.datavalidation import DataValidation
        
        ws = self.wb.create_sheet('RTI')
        ws.sheet_properties.tabColor = COLORS['partner_1']
        
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        row = 2
        
        ws.merge_cells(f'B{row}:H{row}')
        ws[f'B{row}'] = 'CONTRIBUTI PER AZIENDA RTI'
        ws[f'B{row}'].font = TITLE_FONT
        row += 2
        
        # Quote and amounts
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = 'RIPARTIZIONE ECONOMICA'
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        quote_headers = ['Azienda', 'Quota %', 'Importo €']
        for col, header in enumerate(quote_headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        row += 1
        
        quote_start = row
        
        # Create data validation for company names
        company_list = ','.join(self.rti_companies)
        company_dv = DataValidation(type='list', formula1=f'"{company_list}"', allow_blank=False)
        company_dv.error = 'Selezionare un\'azienda dalla lista'
        company_dv.errorTitle = 'Azienda non valida'
        ws.add_data_validation(company_dv)
        
        for company in self.rti_companies:
            quota = self.rti_quotas.get(company, 0)
            
            company_cell = ws.cell(row=row, column=2, value=company)
            company_cell.font = Font(bold=True, color=self.company_colors.get(company, COLORS['dark']))
            company_cell.border = THIN_BORDER
            company_dv.add(company_cell)
            
            # Quota as editable input
            quota_cell = ws.cell(row=row, column=3, value=quota / 100)
            quota_cell.number_format = '0.0%'
            quota_cell.fill = INPUT_FILL
            quota_cell.border = THIN_BORDER
            
            # Importo as FORMULA referencing Economico sheet
            importo_cell = ws.cell(row=row, column=4)
            importo_cell.value = f'=Economico!C{self.econ_prezzo_mio_row}*C{row}'
            importo_cell.number_format = '€ #,##0.00'
            importo_cell.fill = FORMULA_FILL
            importo_cell.border = THIN_BORDER
            
            row += 1
        
        # Total row with formulas
        ws.cell(row=row, column=2, value='TOTALE').font = Font(bold=True)
        ws.cell(row=row, column=2).border = MEDIUM_BORDER
        ws.cell(row=row, column=2).fill = LIGHT_FILL
        ws.cell(row=row, column=3, value=f'=SUM(C{quote_start}:C{row-1})').number_format = '0.0%'
        ws.cell(row=row, column=3).border = MEDIUM_BORDER
        ws.cell(row=row, column=3).fill = LIGHT_FILL
        ws.cell(row=row, column=4, value=f'=SUM(D{quote_start}:D{row-1})').number_format = '€ #,##0.00'
        ws.cell(row=row, column=4).border = MEDIUM_BORDER
        ws.cell(row=row, column=4).fill = LIGHT_FILL
        row += 3
        
        # References
        ws.merge_cells(f'B{row}:G{row}')
        ws[f'B{row}'] = 'REFERENZE (Max Peso Gara assegnato all\'azienda responsabile)'
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        reference_reqs = [r for r in self.lot_config.get('reqs', []) if r.get('type') == 'reference']
        
        if reference_reqs:
            ref_headers = ['ID', 'Requisito', 'Azienda Responsabile', 'Peso Gara (Max)']
            for col, header in enumerate(ref_headers, start=2):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER
                cell.border = THIN_BORDER
            row += 1
            
            for req in reference_reqs:
                req_id = req.get('id', '')
                tech_input = self.tech_inputs_full.get(req_id, {})
                assigned = tech_input.get('assigned_company', '') or 'Lutech'
                gara_weight = req.get('gara_weight', 0)
                
                id_cell = ws.cell(row=row, column=2, value=req_id)
                id_cell.font = Font(size=9, color=COLORS['muted'])
                id_cell.border = THIN_BORDER
                
                label_cell = ws.cell(row=row, column=3, value=req.get('label', '')[:40])
                label_cell.alignment = WRAP
                label_cell.border = THIN_BORDER
                
                assigned_cell = ws.cell(row=row, column=4, value=assigned)
                assigned_cell.font = Font(bold=True, color=self.company_colors.get(assigned, COLORS['dark']))
                assigned_cell.fill = INPUT_FILL
                assigned_cell.border = THIN_BORDER
                company_dv.add(assigned_cell)
                
                weight_cell = ws.cell(row=row, column=5, value=gara_weight)
                weight_cell.number_format = '0.00'
                weight_cell.font = Font(bold=True)
                weight_cell.border = THIN_BORDER
                
                row += 1
            row += 1
        
        # Projects
        ws.merge_cells(f'B{row}:G{row}')
        ws[f'B{row}'] = 'PROGETTI (Max Peso Gara assegnato all\'azienda responsabile)'
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        project_reqs = [r for r in self.lot_config.get('reqs', []) if r.get('type') == 'project']
        
        if project_reqs:
            proj_headers = ['ID', 'Requisito', 'Azienda Responsabile', 'Peso Gara (Max)']
            for col, header in enumerate(proj_headers, start=2):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER
                cell.border = THIN_BORDER
            row += 1
            
            for req in project_reqs:
                req_id = req.get('id', '')
                tech_input = self.tech_inputs_full.get(req_id, {})
                assigned = tech_input.get('assigned_company', '') or 'Lutech'
                gara_weight = req.get('gara_weight', 0)
                
                id_cell = ws.cell(row=row, column=2, value=req_id)
                id_cell.font = Font(size=9, color=COLORS['muted'])
                id_cell.border = THIN_BORDER
                
                label_cell = ws.cell(row=row, column=3, value=req.get('label', '')[:40])
                label_cell.alignment = WRAP
                label_cell.border = THIN_BORDER
                
                assigned_cell = ws.cell(row=row, column=4, value=assigned)
                assigned_cell.font = Font(bold=True, color=self.company_colors.get(assigned, COLORS['dark']))
                assigned_cell.fill = INPUT_FILL
                assigned_cell.border = THIN_BORDER
                company_dv.add(assigned_cell)
                
                weight_cell = ws.cell(row=row, column=5, value=gara_weight)
                weight_cell.number_format = '0.00'
                weight_cell.font = Font(bold=True)
                weight_cell.border = THIN_BORDER
                
                row += 1
            row += 1
        
        row += 1
        
        # Company summary
        ws.merge_cells(f'B{row}:H{row}')
        ws[f'B{row}'] = 'RIEPILOGO CONTRIBUTI PER AZIENDA'
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        summary_headers = ['Azienda', 'Cert. Prof. (prop.)', 'Referenze (max)', 'Progetti (max)', 'TOTALE', '% Contributo']
        for col, header in enumerate(summary_headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        row += 1
        
        summary_start = row
        company_contributions = {company: {'resource': 0, 'reference': 0, 'project': 0} for company in self.rti_companies}
        
        for req in self.lot_config.get('reqs', []):
            req_id = req.get('id', '')
            req_type = req.get('type', 'resource')
            gara_weight = req.get('gara_weight', 0)
            tech_input = self.tech_inputs_full.get(req_id, {})
            
            if req_type == 'resource':
                cert_company_counts = tech_input.get('cert_company_counts', {})
                total_certs = 0
                company_certs = {c: 0 for c in self.rti_companies}
                
                for cert_type, counts in cert_company_counts.items():
                    for company, count in counts.items():
                        if company in company_certs:
                            company_certs[company] += count
                            total_certs += count
                
                if total_certs > 0:
                    weighted_score = self.weighted_scores.get(req_id, 0)
                    for company in self.rti_companies:
                        proportion = company_certs[company] / total_certs
                        company_contributions[company]['resource'] += weighted_score * proportion
                else:
                    company_contributions['Lutech']['resource'] += self.weighted_scores.get(req_id, 0)
                    
            elif req_type in ['reference', 'project']:
                assigned = tech_input.get('assigned_company', '') or 'Lutech'
                if assigned in company_contributions:
                    company_contributions[assigned][req_type] += gara_weight
        
        grand_total = 0
        for company in self.rti_companies:
            contrib = company_contributions[company]
            total = contrib['resource'] + contrib['reference'] + contrib['project']
            grand_total += total
        
        for company in self.rti_companies:
            contrib = company_contributions[company]
            total = contrib['resource'] + contrib['reference'] + contrib['project']
            pct = total / grand_total if grand_total > 0 else 0
            
            ws.cell(row=row, column=2, value=company).font = Font(bold=True, color=self.company_colors.get(company, COLORS['dark']))
            ws.cell(row=row, column=3, value=contrib['resource']).number_format = '0.00'
            ws.cell(row=row, column=4, value=contrib['reference']).number_format = '0.00'
            ws.cell(row=row, column=5, value=contrib['project']).number_format = '0.00'
            ws.cell(row=row, column=6, value=total).number_format = '0.00'
            ws.cell(row=row, column=6).font = Font(bold=True)
            ws.cell(row=row, column=7, value=pct).number_format = '0.0%'
            
            for c in range(2, 8):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1
        
        ws.cell(row=row, column=2, value='TOTALE').font = Font(bold=True)
        for col in range(3, 7):
            ws.cell(row=row, column=col, value=f'=SUM({get_column_letter(col)}{summary_start}:{get_column_letter(col)}{row-1})').number_format = '0.00'
        ws.cell(row=row, column=7, value=f'=SUM(G{summary_start}:G{row-1})').number_format = '0.0%'
        for c in range(2, 8):
            ws.cell(row=row, column=c).fill = LIGHT_FILL
            ws.cell(row=row, column=c).border = MEDIUM_BORDER
        
        rule = DataBarRule(
            start_type='num', start_value=0,
            end_type='num', end_value=1,
            color=COLORS['primary_light']
        )
        ws.conditional_formatting.add(f'G{summary_start}:G{row-1}', rule)

    def _create_config_sheet(self):
        """Create the configuration sheet"""
        ws = self.wb.create_sheet('Config')
        ws.sheet_properties.tabColor = COLORS['muted']
        
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        
        row = 2
        
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = 'CONFIGURAZIONE LOTTO'
        ws[f'B{row}'].font = TITLE_FONT
        row += 2
        
        config_data = [
            ('Nome Lotto', self.lot_key),
            ('Base d\'Asta', f'€ {self.base_amount:,.2f}'),
            ('Max Punteggio Tecnico', self.max_tech_score),
            ('Max Punteggio Economico', self.max_econ_score),
            ('Alpha (α)', self.alpha),
            ('Formula Economica', self.lot_config.get('economic_formula', 'interp_alpha')),
            ('RTI Abilitato', 'Sì' if self.is_rti else 'No'),
        ]
        
        for label, value in config_data:
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = LABEL_FONT
            ws[f'C{row}'] = value
            ws[f'C{row}'].border = THIN_BORDER
            row += 1
        
        if self.is_rti:
            ws[f'B{row}'] = 'Partner RTI'
            ws[f'B{row}'].font = LABEL_FONT
            ws[f'C{row}'] = ', '.join(self.lot_config.get('rti_companies', []))
            ws[f'C{row}'].border = THIN_BORDER
            row += 1
        
        row += 2
        
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = 'NAMED RANGES DEFINITI'
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        ws[f'B{row}'] = 'Nome'
        ws[f'B{row}'].font = HEADER_FONT
        ws[f'B{row}'].fill = HEADER_FILL
        ws[f'C{row}'] = 'Riferimento'
        ws[f'C{row}'].font = HEADER_FONT
        ws[f'C{row}'].fill = HEADER_FILL
        row += 1
        
        for name, ref in self.named_ranges.items():
            ws[f'B{row}'] = name
            ws[f'B{row}'].font = Font(name='Consolas', size=10)
            ws[f'C{row}'] = ref
            ws[f'C{row}'].font = FORMULA_FONT
            row += 1
        
        row += 2
        
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = 'METADATI'
        ws[f'B{row}'].font = SECTION_FONT
        row += 1
        
        ws[f'B{row}'] = 'Data Generazione'
        ws[f'C{row}'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        row += 1
        
        ws[f'B{row}'] = 'Versione App'
        ws[f'C{row}'] = '2.0.0'

    def _create_named_ranges(self):
        """Create Excel named ranges for cross-sheet references"""
        for name, ref in self.named_ranges.items():
            defn = DefinedName(name, attr_text=ref)
            self.wb.defined_names[name] = defn


def generate_excel_report(
    lot_key: str,
    lot_config: Dict[str, Any],
    base_amount: float,
    my_discount: float,
    competitor_discount: float,
    technical_score: float,
    economic_score: float,
    total_score: float,
    details: Dict[str, float],
    weighted_scores: Dict[str, float],
    category_scores: Dict[str, float],
    max_tech_score: float,
    max_econ_score: float,
    alpha: float,
    win_probability: float,
    tech_inputs_full: Optional[Dict[str, Any]] = None,
    rti_quotas: Optional[Dict[str, float]] = None,
) -> io.BytesIO:
    """Generate Excel report"""
    generator = ExcelReportGenerator(
        lot_key=lot_key,
        lot_config=lot_config,
        base_amount=base_amount,
        my_discount=my_discount,
        competitor_discount=competitor_discount,
        technical_score=technical_score,
        economic_score=economic_score,
        total_score=total_score,
        details=details,
        weighted_scores=weighted_scores,
        category_scores=category_scores,
        max_tech_score=max_tech_score,
        max_econ_score=max_econ_score,
        alpha=alpha,
        win_probability=win_probability,
        tech_inputs_full=tech_inputs_full,
        rti_quotas=rti_quotas,
    )
    
    return generator.generate()
