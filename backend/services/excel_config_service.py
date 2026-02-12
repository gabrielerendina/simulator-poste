"""
Excel Configuration Service

Provides functionality to generate Excel templates for lot configuration
and parse uploaded Excel files to create LotConfig objects.
"""

import io
import logging
from typing import List, Dict, Any, Tuple, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Sheet names
SHEET_LOTTO = "Lotto"
SHEET_CERT_AZIENDALI = "Cert_Aziendali"
SHEET_CERT_PROFESSIONALI = "Cert_Professionali"
SHEET_REFERENZE_PROGETTI = "Referenze_Progetti"
SHEET_CRITERI = "Criteri"
SHEET_VOCI_TABELLARI = "Voci_Tabellari"
SHEET_MD_CERT_AZ = "_MD_CertAziendali"
SHEET_MD_CERT_PROF = "_MD_CertProfessionali"
SHEET_MD_FORMULE = "_MD_Formule"
SHEET_MD_TIPI = "_MD_Tipi"

# Styles
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
MD_HEADER_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
MD_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class ExcelConfigService:
    """Service for Excel template generation and parsing."""

    @staticmethod
    def _find_sheet(wb: Workbook, expected_name: str) -> Optional[Worksheet]:
        """Find sheet by name with case-insensitive matching."""
        expected_lower = expected_name.lower()
        for name in wb.sheetnames:
            if name.lower() == expected_lower:
                return wb[name]
        return None

    @staticmethod
    def export_lot_config(lot_config: Dict[str, Any], master_data: Dict[str, Any]) -> io.BytesIO:
        """
        Export a lot configuration to a populated Excel file.
        
        Args:
            lot_config: The lot configuration dict to export
            master_data: Master data dict with company_certs, prof_certs, economic_formulas
            
        Returns:
            BytesIO object containing the Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create master data sheets first (for named ranges)
        ExcelConfigService._create_md_sheets(wb, master_data)
        
        # Create and populate input sheets
        ExcelConfigService._export_lotto_sheet(wb, lot_config, master_data)
        ExcelConfigService._export_cert_aziendali_sheet(wb, lot_config)
        ExcelConfigService._export_cert_professionali_sheet(wb, lot_config)
        ExcelConfigService._export_referenze_progetti_sheet(wb, lot_config)
        ExcelConfigService._export_criteri_sheet(wb, lot_config)
        ExcelConfigService._export_voci_tabellari_sheet(wb, lot_config)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _export_lotto_sheet(wb: Workbook, lot_config: Dict[str, Any], master_data: Dict[str, Any]):
        """Create and populate Lotto sheet with lot config data."""
        ws = wb.create_sheet(SHEET_LOTTO, 0)
        headers = ["Nome", "Importo Base", "Alpha", "Formula Economica"]
        ws.append(headers)
        ExcelConfigService._style_header(ws, 1, len(headers))
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        
        # Add data validation for formula_economica
        formulas = master_data.get("economic_formulas", [])
        if formulas:
            dv = DataValidation(type="list", formula1="MD_Formule", allow_blank=False)
            ws.add_data_validation(dv)
            dv.add("D2:D100")
        
        # Populate with lot data
        ws.append([
            lot_config.get("name", ""),
            lot_config.get("base_amount", 0),
            lot_config.get("alpha", 0.3),
            lot_config.get("economic_formula", "interp_alpha")
        ])

    @staticmethod
    def _export_cert_aziendali_sheet(wb: Workbook, lot_config: Dict[str, Any]):
        """Create and populate company certifications sheet."""
        ws = wb.create_sheet(SHEET_CERT_AZIENDALI, 1)
        headers = ["Nome Certificazione", "Punti", "Punti Parziale RTI", "Peso Gara"]
        ws.append(headers)
        ExcelConfigService._style_header(ws, 1, len(headers))
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        
        # Populate with company certs
        for cert in lot_config.get("company_certs", []):
            ws.append([
                cert.get("label", ""),
                cert.get("points", 0),
                cert.get("points_partial", 0),
                cert.get("gara_weight", 0)
            ])

    @staticmethod
    def _export_cert_professionali_sheet(wb: Workbook, lot_config: Dict[str, Any]):
        """Create and populate professional certifications sheet."""
        ws = wb.create_sheet(SHEET_CERT_PROFESSIONALI, 2)
        headers = [
            "Codice", "Nome", "Peso Gara", "Max Punti Raw", 
            "Risorse Richieste", "Cert per Risorsa", "Certificazioni (sep. ;)"
        ]
        ws.append(headers)
        ExcelConfigService._style_header(ws, 1, len(headers))
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 60
        
        # Populate resource-type requirements
        for req in lot_config.get("reqs", []):
            if req.get("type") == "resource":
                certs_str = ";".join(req.get("selected_prof_certs", []))
                ws.append([
                    req.get("id", ""),
                    req.get("label", ""),
                    req.get("gara_weight", 0),
                    req.get("max_points", 0),
                    req.get("prof_R", 0),
                    req.get("prof_C", 0),
                    certs_str
                ])

    @staticmethod
    def _export_referenze_progetti_sheet(wb: Workbook, lot_config: Dict[str, Any]):
        """Create and populate references/projects sheet."""
        ws = wb.create_sheet(SHEET_REFERENZE_PROGETTI, 3)
        headers = ["Codice", "Tipo", "Nome", "Peso Gara", "Attestazione Cliente"]
        ws.append(headers)
        ExcelConfigService._style_header(ws, 1, len(headers))
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        
        # Data validation for tipo
        dv = DataValidation(type="list", formula1="MD_Tipi", allow_blank=False)
        ws.add_data_validation(dv)
        dv.add("B2:B100")
        
        # Populate reference/project type requirements
        for req in lot_config.get("reqs", []):
            if req.get("type") in ("reference", "project"):
                ws.append([
                    req.get("id", ""),
                    req.get("type", "reference"),
                    req.get("label", ""),
                    req.get("gara_weight", 0),
                    req.get("attestazione_score", 0)
                ])

    @staticmethod
    def _export_criteri_sheet(wb: Workbook, lot_config: Dict[str, Any]):
        """Create and populate criteria sheet."""
        ws = wb.create_sheet(SHEET_CRITERI, 4)
        headers = [
            "Codice Requisito", "Nome Criterio", "Peso",
            "Assente/Inadeguato", "Parzialmente Adeguato", "Adeguato", 
            "Più che Adeguato", "Ottimo"
        ]
        ws.append(headers)
        ExcelConfigService._style_header(ws, 1, len(headers))
        
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 10
        
        # Populate criteria from reference/project requirements
        for req in lot_config.get("reqs", []):
            if req.get("type") in ("reference", "project"):
                for sub_req in req.get("sub_reqs", []):
                    judgement = sub_req.get("judgement_levels", {})
                    ws.append([
                        req.get("id", ""),
                        sub_req.get("label", ""),
                        sub_req.get("weight", 1.0),
                        judgement.get("assente_inadeguato", 0),
                        judgement.get("parzialmente_adeguato", 2),
                        judgement.get("adeguato", 3),
                        judgement.get("piu_che_adeguato", 4),
                        judgement.get("ottimo", 5)
                    ])

    @staticmethod
    def _export_voci_tabellari_sheet(wb: Workbook, lot_config: Dict[str, Any]):
        """Create and populate tabular metrics sheet."""
        ws = wb.create_sheet(SHEET_VOCI_TABELLARI, 5)
        headers = ["Codice Requisito", "ID Voce", "Nome Voce", "Min Score", "Max Score"]
        ws.append(headers)
        ExcelConfigService._style_header(ws, 1, len(headers))
        
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        # Populate custom metrics from reference/project requirements
        for req in lot_config.get("reqs", []):
            if req.get("type") in ("reference", "project"):
                for metric in req.get("custom_metrics", []):
                    ws.append([
                        req.get("id", ""),
                        metric.get("id", ""),
                        metric.get("label", ""),
                        metric.get("min_score", 0),
                        metric.get("max_score", 5)
                    ])

    @staticmethod
    def generate_template(master_data: Dict[str, Any]) -> io.BytesIO:
        """
        Generate an Excel template for lot configuration.
        
        Args:
            master_data: Master data dict with company_certs, prof_certs, economic_formulas
            
        Returns:
            BytesIO object containing the Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create master data sheets first (for named ranges)
        ExcelConfigService._create_md_sheets(wb, master_data)
        
        # Create input sheets
        ExcelConfigService._create_lotto_sheet(wb, master_data)
        ExcelConfigService._create_cert_aziendali_sheet(wb, master_data)
        ExcelConfigService._create_cert_professionali_sheet(wb, master_data)
        ExcelConfigService._create_referenze_progetti_sheet(wb)
        ExcelConfigService._create_criteri_sheet(wb)
        ExcelConfigService._create_voci_tabellari_sheet(wb)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _style_header(ws: Worksheet, row: int, cols: int, fill=HEADER_FILL, font=HEADER_FONT):
        """Apply header styling to a row."""
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    @staticmethod
    def _create_md_sheets(wb: Workbook, master_data: Dict[str, Any]):
        """Create master data reference sheets."""
        # _MD_CertAziendali
        ws = wb.create_sheet(SHEET_MD_CERT_AZ)
        ws.append(["Certificazioni Aziendali Disponibili"])
        ExcelConfigService._style_header(ws, 1, 1, MD_HEADER_FILL, HEADER_FONT)
        company_certs = master_data.get("company_certs", [])
        for cert in company_certs:
            ws.append([cert])
        ws.column_dimensions['A'].width = 35
        # Create named range for dropdown
        if company_certs:
            wb.create_named_range("MD_CertAziendali", ws, f"$A$2:$A${len(company_certs) + 1}")
        # Protect sheet
        ws.protection.sheet = True
        ws.protection.password = 'readonly'
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = MD_FILL
                cell.protection = Protection(locked=True)
        
        # _MD_CertProfessionali
        ws = wb.create_sheet(SHEET_MD_CERT_PROF)
        ws.append(["Certificazioni Professionali Disponibili"])
        ExcelConfigService._style_header(ws, 1, 1, MD_HEADER_FILL, HEADER_FONT)
        prof_certs = master_data.get("prof_certs", [])
        for cert in prof_certs:
            ws.append([cert])
        ws.column_dimensions['A'].width = 40
        if prof_certs:
            wb.create_named_range("MD_CertProfessionali", ws, f"$A$2:$A${len(prof_certs) + 1}")
        ws.protection.sheet = True
        ws.protection.password = 'readonly'
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = MD_FILL
                cell.protection = Protection(locked=True)
        
        # _MD_Formule
        ws = wb.create_sheet(SHEET_MD_FORMULE)
        ws.append(["ID Formula", "Descrizione"])
        ExcelConfigService._style_header(ws, 1, 2, MD_HEADER_FILL, HEADER_FONT)
        formulas = master_data.get("economic_formulas", [])
        for f in formulas:
            ws.append([f.get("id", ""), f.get("label", "")])
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        if formulas:
            wb.create_named_range("MD_Formule", ws, f"$A$2:$A${len(formulas) + 1}")
        ws.protection.sheet = True
        ws.protection.password = 'readonly'
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = MD_FILL
                cell.protection = Protection(locked=True)
        
        # _MD_Tipi
        ws = wb.create_sheet(SHEET_MD_TIPI)
        ws.append(["Tipo Requisito"])
        ExcelConfigService._style_header(ws, 1, 1, MD_HEADER_FILL, HEADER_FONT)
        types = ["reference", "project"]
        for t in types:
            ws.append([t])
        ws.column_dimensions['A'].width = 20
        wb.create_named_range("MD_Tipi", ws, f"$A$2:$A${len(types) + 1}")
        ws.protection.sheet = True
        ws.protection.password = 'readonly'
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = MD_FILL
                cell.protection = Protection(locked=True)

    @staticmethod
    def _create_lotto_sheet(wb: Workbook, master_data: Dict[str, Any]):
        """Create Lotto configuration sheet."""
        ws = wb.create_sheet(SHEET_LOTTO, 0)  # First position
        headers = ["Nome", "Importo Base", "Alpha", "Formula Economica"]
        ws.append(headers)
        ExcelConfigService._style_header(ws, 1, len(headers))
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        
        # Add data validation for formula_economica
        formulas = master_data.get("economic_formulas", [])
        if formulas:
            dv = DataValidation(
                type="list",
                formula1="MD_Formule",
                allow_blank=False
            )
            dv.error = "Seleziona una formula dalla lista"
            dv.errorTitle = "Formula non valida"
            ws.add_data_validation(dv)
            dv.add("D2:D100")
        
        # Add sample row with example values
        ws.append(["Nome del Lotto", 1000000, 0.3, "interp_alpha"])

    @staticmethod
    def _create_cert_aziendali_sheet(wb: Workbook, master_data: Dict[str, Any]):
        """Create company certifications sheet."""
        ws = wb.create_sheet(SHEET_CERT_AZIENDALI, 1)
        headers = ["Nome Certificazione", "Punti", "Punti Parziale RTI", "Peso Gara"]
        ws.append(headers)
        ExcelConfigService._style_header(ws, 1, len(headers))
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        
        # Data validation for nome
        company_certs = master_data.get("company_certs", [])
        if company_certs:
            dv = DataValidation(
                type="list",
                formula1="MD_CertAziendali",
                allow_blank=True
            )
            dv.error = "Seleziona una certificazione dalla lista"
            dv.errorTitle = "Certificazione non valida"
            ws.add_data_validation(dv)
            dv.add("A2:A100")

    @staticmethod
    def _create_cert_professionali_sheet(wb: Workbook, master_data: Dict[str, Any]):
        """Create professional certifications / resource requirements sheet."""
        ws = wb.create_sheet(SHEET_CERT_PROFESSIONALI, 2)
        headers = [
            "Codice", "Nome", "Peso Gara", "Max Punti Raw", 
            "Risorse Richieste", "Cert per Risorsa", "Certificazioni (sep. ;)"
        ]
        ws.append(headers)
        ExcelConfigService._style_header(ws, 1, len(headers))
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 60
        
        # Add comment for certificazioni column
        ws["G1"].comment = None  # Clear any existing
        # Note: certificazioni should be semicolon-separated list from MD_CertProfessionali

    @staticmethod
    def _create_referenze_progetti_sheet(wb: Workbook):
        """Create references/projects requirements sheet."""
        ws = wb.create_sheet(SHEET_REFERENZE_PROGETTI, 3)
        headers = ["Codice", "Tipo", "Nome", "Peso Gara", "Attestazione Cliente"]
        ws.append(headers)
        ExcelConfigService._style_header(ws, 1, len(headers))
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        
        # Data validation for tipo
        dv = DataValidation(
            type="list",
            formula1="MD_Tipi",
            allow_blank=False
        )
        dv.error = "Seleziona 'reference' o 'project'"
        dv.errorTitle = "Tipo non valido"
        ws.add_data_validation(dv)
        dv.add("B2:B100")

    @staticmethod
    def _create_criteri_sheet(wb: Workbook):
        """Create criteria sheet for references/projects with judgement levels."""
        ws = wb.create_sheet(SHEET_CRITERI, 4)
        headers = [
            "Codice Requisito", "Nome Criterio", "Peso",
            "Assente/Inadeguato", "Parzialmente Adeguato", "Adeguato", 
            "Più che Adeguato", "Ottimo"
        ]
        ws.append(headers)
        ExcelConfigService._style_header(ws, 1, len(headers))
        
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 10
        
        # Add example row with default judgement values (0, 2, 3, 4, 5)
        ws.append(["REF_001", "Criterio esempio", 1.0, 0, 2, 3, 4, 5])

    @staticmethod
    def _create_voci_tabellari_sheet(wb: Workbook):
        """Create tabular metrics sheet."""
        ws = wb.create_sheet(SHEET_VOCI_TABELLARI, 5)
        headers = ["Codice Requisito", "ID Voce", "Nome Voce", "Min Score", "Max Score"]
        ws.append(headers)
        ExcelConfigService._style_header(ws, 1, len(headers))
        
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        # Add example row
        ws.append(["REF_001", "M1", "Voce tabellare esempio", 0, 5])

    @staticmethod
    def parse_upload(file_content: bytes, master_data: Dict[str, Any]) -> Tuple[Dict[str, Any], List[str]]:
        """
        Parse an uploaded Excel file and build a LotConfig.
        
        Args:
            file_content: Raw bytes of the uploaded Excel file
            master_data: Master data for validation
            
        Returns:
            Tuple of (lot_config_dict, warnings_list)
        """
        warnings = []
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        
        logger.info(f"Excel upload: found sheets {wb.sheetnames}")
        
        # Debug: log sheet details
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"  Sheet '{sheet_name}': max_row={ws.max_row}, max_col={ws.max_column}")
        
        # Parse Lotto sheet
        lot_config = ExcelConfigService._parse_lotto_sheet(wb, warnings)
        if not lot_config:
            return None, [f"Foglio 'Lotto' non trovato o vuoto. Fogli disponibili: {wb.sheetnames}"]
        
        logger.info(f"Parsed Lotto: name={lot_config.get('name')}")
        
        # Parse Cert_Aziendali
        company_certs = ExcelConfigService._parse_cert_aziendali_sheet(wb, master_data, warnings)
        lot_config["company_certs"] = company_certs
        logger.info(f"Parsed Cert_Aziendali: {len(company_certs)} items")
        
        # Parse requirements (reqs list will contain all requirement types)
        reqs = []
        
        # Parse Cert_Professionali (resource type requirements)
        resource_reqs = ExcelConfigService._parse_cert_professionali_sheet(wb, master_data, warnings)
        reqs.extend(resource_reqs)
        logger.info(f"Parsed Cert_Professionali: {len(resource_reqs)} requirements")
        
        # Parse Referenze_Progetti
        ref_proj_reqs = ExcelConfigService._parse_referenze_progetti_sheet(wb, warnings)
        logger.info(f"Parsed Referenze_Progetti: {len(ref_proj_reqs)} requirements")
        
        # Parse Criteri and attach to requirements
        criteri_map = ExcelConfigService._parse_criteri_sheet(wb, warnings)
        logger.info(f"Parsed Criteri: {len(criteri_map)} requirement codes")
        
        # Parse Voci_Tabellari and attach to requirements
        voci_map = ExcelConfigService._parse_voci_tabellari_sheet(wb, warnings)
        logger.info(f"Parsed Voci_Tabellari: {len(voci_map)} requirement codes")
        
        # Attach criteri and voci to ref/proj requirements
        for req in ref_proj_reqs:
            req_code = req["id"]
            if req_code in criteri_map:
                req["sub_reqs"] = criteri_map[req_code]
                logger.debug(f"Attached {len(criteri_map[req_code])} criteria to {req_code}")
            if req_code in voci_map:
                req["custom_metrics"] = voci_map[req_code]
        
        reqs.extend(ref_proj_reqs)
        lot_config["reqs"] = reqs
        
        # Add warnings for empty sections
        if not resource_reqs and ExcelConfigService._find_sheet(wb, SHEET_CERT_PROFESSIONALI):
            warnings.append("Foglio 'Cert_Professionali' trovato ma nessun requisito valido")
        if not ref_proj_reqs and ExcelConfigService._find_sheet(wb, SHEET_REFERENZE_PROGETTI):
            warnings.append("Foglio 'Referenze_Progetti' trovato ma nessun requisito valido")
        
        # Calculate max_raw_score
        max_raw = sum(cert.get("points", 0) for cert in company_certs)
        max_raw += sum(req.get("max_points", 0) for req in reqs)
        lot_config["max_raw_score"] = max_raw
        
        # Calculate max_tech_score from gara_weights (override hardcoded default)
        max_tech = sum(cert.get("gara_weight", 0) for cert in company_certs)
        max_tech += sum(req.get("gara_weight", 0) for req in reqs)
        if max_tech > 0:
            lot_config["max_tech_score"] = max_tech
            lot_config["max_econ_score"] = 100.0 - max_tech
        
        logger.info(f"Import complete: {len(reqs)} total requirements, max_raw={max_raw}, max_tech={max_tech}")
        
        return lot_config, warnings

    @staticmethod
    def _parse_lotto_sheet(wb: Workbook, warnings: List[str]) -> Optional[Dict[str, Any]]:
        """Parse Lotto sheet."""
        ws = ExcelConfigService._find_sheet(wb, SHEET_LOTTO)
        if ws is None:
            logger.warning(f"Sheet '{SHEET_LOTTO}' not found. Available: {wb.sheetnames}")
            return None
        
        # Find header row
        headers = [cell.value for cell in ws[1]]
        logger.debug(f"Lotto headers: {headers}")
        
        # Read first data row
        if ws.max_row < 2:
            logger.warning(f"Lotto sheet has no data rows (max_row={ws.max_row})")
            return None
        
        row = ws[2]
        values = [cell.value for cell in row]
        logger.debug(f"Lotto data row: {values}")
        
        config = {
            "name": str(values[0]).strip() if values[0] else "",
            "base_amount": float(values[1]) if len(values) > 1 and values[1] else 0.0,
            "alpha": float(values[2]) if len(values) > 2 and values[2] else 0.3,
            "economic_formula": str(values[3]).strip() if len(values) > 3 and values[3] else "interp_alpha",
            "max_tech_score": 60.0,
            "max_econ_score": 40.0,
        }
        
        logger.info(f"Parsed Lotto config: name='{config['name']}', base={config['base_amount']}")
        
        if not config["name"]:
            warnings.append("Nome lotto mancante, utilizza valore default")
            config["name"] = "Lotto_Import"
        
        return config

    @staticmethod
    def _parse_cert_aziendali_sheet(
        wb: Workbook, 
        master_data: Dict[str, Any], 
        warnings: List[str]
    ) -> List[Dict[str, Any]]:
        """Parse company certifications sheet."""
        ws = ExcelConfigService._find_sheet(wb, SHEET_CERT_AZIENDALI)
        if ws is None:
            logger.info(f"Sheet '{SHEET_CERT_AZIENDALI}' not found, skipping")
            return []
        valid_certs = set(master_data.get("company_certs", []))
        result = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            
            nome = str(row[0]).strip()
            punti = float(row[1]) if row[1] else 0.0
            punti_parziale = float(row[2]) if len(row) > 2 and row[2] else 0.0
            peso_gara = float(row[3]) if len(row) > 3 and row[3] else 0.0
            
            if nome not in valid_certs:
                warnings.append(f"Riga {row_num} Cert_Aziendali: '{nome}' non in master data, saltata")
                continue
            
            result.append({
                "label": nome,
                "points": punti,
                "points_partial": punti_parziale,
                "gara_weight": peso_gara
            })
        
        return result

    @staticmethod
    def _parse_cert_professionali_sheet(
        wb: Workbook,
        master_data: Dict[str, Any],
        warnings: List[str]
    ) -> List[Dict[str, Any]]:
        """Parse professional certifications (resource requirements) sheet."""
        ws = ExcelConfigService._find_sheet(wb, SHEET_CERT_PROFESSIONALI)
        if ws is None:
            logger.info(f"Sheet '{SHEET_CERT_PROFESSIONALI}' not found, skipping")
            return []
        
        # Log headers for debugging
        headers = [cell.value for cell in ws[1]]
        logger.info(f"Cert_Professionali: headers={headers}, max_row={ws.max_row}")
        
        valid_prof_certs = set(master_data.get("prof_certs", []))
        result = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                logger.debug(f"Row {row_num}: empty row tuple")
                continue
            if not row[0]:
                logger.debug(f"Row {row_num}: first cell is empty, row={row}")
                continue
            
            codice = str(row[0]).strip()
            nome = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            peso_gara = float(row[2]) if len(row) > 2 and row[2] else 0.0
            max_punti = float(row[3]) if len(row) > 3 and row[3] else 0.0
            risorse_rich = int(row[4]) if len(row) > 4 and row[4] else 0
            cert_per_risorsa = int(row[5]) if len(row) > 5 and row[5] else 0
            certificazioni_raw = str(row[6]).strip() if len(row) > 6 and row[6] else ""
            
            logger.info(f"Cert_Professionali row {row_num}: codice={codice}, nome={nome}")
            
            # Parse semicolon-separated certifications
            selected_certs = []
            if certificazioni_raw:
                for cert in certificazioni_raw.split(";"):
                    cert = cert.strip()
                    if cert:
                        if cert not in valid_prof_certs:
                            warnings.append(
                                f"Riga {row_num} Cert_Professionali: certificazione '{cert}' non in master data"
                            )
                        selected_certs.append(cert)
            
            req = {
                "id": codice,
                "label": nome,
                "type": "resource",
                "max_points": max_punti,
                "max_points_manual": max_punti > 0,  # Use manual max if explicitly set
                "gara_weight": peso_gara,
                "prof_R": risorse_rich,
                "prof_C": cert_per_risorsa,
                "selected_prof_certs": selected_certs,
            }
            result.append(req)
        
        return result

    @staticmethod
    def _parse_referenze_progetti_sheet(wb: Workbook, warnings: List[str]) -> List[Dict[str, Any]]:
        """Parse references/projects requirements sheet."""
        ws = ExcelConfigService._find_sheet(wb, SHEET_REFERENZE_PROGETTI)
        if ws is None:
            logger.info(f"Sheet '{SHEET_REFERENZE_PROGETTI}' not found, skipping")
            return []
        
        # Log headers for debugging
        headers = [cell.value for cell in ws[1]]
        logger.info(f"Referenze_Progetti: headers={headers}, max_row={ws.max_row}")
        
        valid_types = {"reference", "project"}
        result = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                logger.debug(f"Row {row_num}: empty row tuple")
                continue
            if not row[0]:
                logger.debug(f"Row {row_num}: first cell is empty, row={row}")
                continue
            
            codice = str(row[0]).strip()
            tipo = str(row[1]).strip().lower() if len(row) > 1 and row[1] else "reference"
            nome = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            peso_gara = float(row[3]) if len(row) > 3 and row[3] else 0.0
            attestazione = float(row[4]) if len(row) > 4 and row[4] else 0.0
            
            if tipo not in valid_types:
                warnings.append(f"Riga {row_num} Referenze_Progetti: tipo '{tipo}' non valido, uso 'reference'")
                tipo = "reference"
            
            req = {
                "id": codice,
                "label": nome,
                "type": tipo,
                "max_points": 0.0,  # Will be calculated from criteria
                "gara_weight": peso_gara,
                "attestazione_score": attestazione,
                "sub_reqs": [],
                "custom_metrics": [],
            }
            result.append(req)
        
        return result

    @staticmethod
    def _parse_criteri_sheet(wb: Workbook, warnings: List[str]) -> Dict[str, List[Dict[str, Any]]]:
        """Parse criteria sheet and return map of req_code -> criteria list with judgement levels."""
        ws = ExcelConfigService._find_sheet(wb, SHEET_CRITERI)
        if ws is None:
            logger.info(f"Sheet '{SHEET_CRITERI}' not found, skipping")
            return {}
        
        result = {}
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            
            codice_req = str(row[0]).strip()
            criterio_nome = str(row[1]).strip() if row[1] else ""
            peso = float(row[2]) if row[2] else 1.0
            
            # Parse judgement levels (columns D through H)
            # Default values: 0, 2, 3, 4, 5
            assente = float(row[3]) if len(row) > 3 and row[3] is not None else 0
            parziale = float(row[4]) if len(row) > 4 and row[4] is not None else 2
            adeguato = float(row[5]) if len(row) > 5 and row[5] is not None else 3
            piu_che = float(row[6]) if len(row) > 6 and row[6] is not None else 4
            ottimo = float(row[7]) if len(row) > 7 and row[7] is not None else 5
            
            if codice_req not in result:
                result[codice_req] = []
            
            # Use letters (a, b, c, ...) for criterion IDs
            criterion_index = len(result[codice_req])
            criterion_id = chr(97 + criterion_index) if criterion_index < 26 else f"c{criterion_index + 1}"
            result[codice_req].append({
                "id": criterion_id,
                "label": criterio_nome,
                "weight": peso,
                "max_value": ottimo,  # max_value derived from ottimo
                "judgement_levels": {
                    "assente_inadeguato": assente,
                    "parzialmente_adeguato": parziale,
                    "adeguato": adeguato,
                    "piu_che_adeguato": piu_che,
                    "ottimo": ottimo,
                }
            })
        
        return result

    @staticmethod
    def _parse_voci_tabellari_sheet(wb: Workbook, warnings: List[str]) -> Dict[str, List[Dict[str, Any]]]:
        """Parse tabular metrics sheet and return map of req_code -> metrics list."""
        ws = ExcelConfigService._find_sheet(wb, SHEET_VOCI_TABELLARI)
        if ws is None:
            logger.info(f"Sheet '{SHEET_VOCI_TABELLARI}' not found, skipping")
            return {}
        
        result = {}
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            
            codice_req = str(row[0]).strip()
            id_voce = str(row[1]).strip() if row[1] else ""
            nome_voce = str(row[2]).strip() if row[2] else ""
            min_score = float(row[3]) if row[3] else 0.0
            max_score = float(row[4]) if len(row) > 4 and row[4] else 5.0
            
            if codice_req not in result:
                result[codice_req] = []
            
            result[codice_req].append({
                "id": id_voce or f"M{len(result[codice_req]) + 1}",
                "label": nome_voce,
                "min_score": min_score,
                "max_score": max_score,
            })
        
        return result
